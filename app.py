from __future__ import annotations
import calendar
from pathlib import Path

from dataclasses import dataclass
from datetime import datetime, date, timedelta, time
from typing import Dict, Tuple, List, Optional
from copy import copy
from collections import defaultdict
import io
import hashlib
import re

def get_base_dir() -> Path:
    """실행 파일(app.py) 기준 폴더."""
    try:
        return Path(__file__).resolve().parent
    except Exception:
        return Path.cwd()

def get_output_dir() -> Path:
    """output 폴더가 없으면 생성해서 반환."""
    out_dir = get_base_dir() / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _ui_info(title: str, msg: str) -> None:
    """환경에 따라 메시지를 팝업으로 보여주고, 실패하면 콘솔로 출력."""
    try:
        messagebox.showinfo(title, msg)
    except Exception:
        try:
            print(f"[{title}] {msg}")
        except Exception:
            pass

def _ui_error(title: str, msg: str) -> None:
    try:
        messagebox.showerror(title, msg)
    except Exception:
        try:
            print(f"[{title}] {msg}")
        except Exception:
            pass

import tkinter as tk
# 외부 달력(tkcalendar) 없이 동작하도록 고정
DateEntry = None  # type: ignore
HAS_TKCALENDAR = False

from tkinter import ttk, messagebox

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage


# =========================
# Files / Paths
# =========================
# ✅ 항상 app.py가 있는 폴더 기준으로 파일을 찾음 (어디서 실행해도 OK)
BASE_DIR = Path(__file__).resolve().parent
RATES_FILE = BASE_DIR / "rates.xlsx"                 # ✅ 요율 파일
RATES_SHEET = "요율"

# ✅ 템플릿: 기본 + 예외(협성/동진)
TEMPLATE_DEFAULT = BASE_DIR / "invoice_template.xlsx"
TEMPLATE_EXCEPTIONS = {
    "협성": BASE_DIR / "invoice_template_hyupsung.xlsx",
    # 동진 템플릿 실제 파일명에 맞게 유지 (기존 코드 기준)
    "동진": BASE_DIR / "invoice_template_dongjin.xlsx",
}


def resolve_template_path(agency_kr: str) -> Path:
    """대리점별 템플릿 경로를 반환. 동진은 파일명이 다를 수 있어 fallback 처리."""
    a = (agency_kr or "").strip()
    if a in TEMPLATE_EXCEPTIONS:
        p = TEMPLATE_EXCEPTIONS[a]
        if p.exists():
            return p
        if a == "동진":
            candidates = [
                BASE_DIR / "invoice_template_dongjin.xlsx",
                BASE_DIR / "invoice_template_동진.xlsx",
                BASE_DIR / "invoice_template_DONGJIN.xlsx",
            ]
            for c in candidates:
                if c.exists():
                    return c
    return TEMPLATE_DEFAULT

OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
IMG_CACHE_DIR = OUTPUT_DIR / "_imgcache"

# 기본 템플릿 룸/토탈 시트
BASE_ROOM_SHEET_DEFAULT = "402"
TOTAL_SHEET_NAME_DEFAULT = "TOTAL INVOICE (HOTEL)"


# =========================
# Data models
# =========================
@dataclass(frozen=True)
class RateRow:
    agency_eng: str
    nightly: int
    late_1201_1400: int
    early_0801_0959: int
    breakfast: int
    lunch: int
    dinner: int


@dataclass(frozen=True)
class CrewStay:
    vessel: str
    room_no: str
    crew_name: str
    nationality: str
    agency_name: str          # UI에 보이는 한글 대리점명
    room_type: str
    checkin: datetime
    checkout: datetime
    meals_by_date: Dict[str, Tuple[bool, bool, bool]]  # "YYYY-MM-DD" -> (B,L,D)
    charge_room: bool         # Twin이면 1명만 True (룸요금 1회만 계산)


# =========================
# Helpers
# =========================
def parse_dt_24h(s: str) -> datetime:
    return datetime.strptime(s.strip(), "%Y-%m-%d %H:%M")


def daterange_inclusive(d1: date, d2: date) -> List[date]:
    if d2 < d1:
        return []
    days = (d2 - d1).days
    return [d1 + timedelta(days=i) for i in range(days + 1)]


def format_date_dd_mmm_yy(d: date) -> str:
    """1-May-26 형식(영문 월 고정)"""
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{d.day}-{months[d.month-1]}-{str(d.year)[-2:]}"

EXCEL_DATE_FMT = '[$-409]d"-"mmm"-"yy;@'
EXCEL_AMOUNT_ZERO_DASH_FMT = '#,##0;[Red](#,##0);-' 


def _to_int(v) -> int:
    """요율 시트에 '숙박 1박' 같은 텍스트가 섞여도 숫자만 안전하게 추출."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    s = s.replace(",", "")
    s = re.sub(r"[^\d\-]", "", s)
    if s in ("", "-", "--"):
        return 0
    return int(s)


def is_twin(room_type: str) -> bool:
    s = (room_type or "").strip().lower()
    raw = (room_type or "").strip()
    return ("twin" in s) or ("트윈" in raw) or ("2인" in raw) or ("두명" in raw)


def roomtype_to_english(room_type: str) -> str:
    s = (room_type or "").strip()
    sl = s.lower()
    if "twin" in sl or "트윈" in s or "2인" in s or "두명" in s:
        return "Twin"
    if "single" in sl or "싱글" in s or "일반" in s:
        return "Single"
    return s

def normalize_multiline(s: str) -> str:
    """'PHL, MMR' 또는 'PHL PHL' 등을 줄바꿈 문자열로 정리."""
    if not s:
        return ""
    t = str(s).strip()
    if not t:
        return ""
    # 콤마/세미콜론 기준 분리
    parts = re.split(r"[;,]", t)
    out = []
    for p in parts:
        p2 = p.strip()
        if not p2:
            continue
        out.append(p2)
    # 콤마가 없었는데 공백으로 여러개 들어온 경우(예: 'PHL PHL')
    if len(out) == 1 and " " in out[0]:
        out = [x for x in out[0].split() if x]
    return "\n".join(out)


def get_accommodation_base_desc(agency_name: str) -> str:
    """대리점별 숙박 기본 문구."""
    a = (agency_name or "").strip()
    # 협성/동진: 조식 포함 문구
    if ("협성" in a) or ("동진" in a):
        return "Accommodation Charge(Breakfast included)"
    # 신양: 세금 포함 문구
    if ("신양" in a) or ("shinyang" in a.lower()):
        return "Accommodation Charge(Include tax)"
    # 그 외 기본
    return "Accommodation Charge"




def _get_sheet_image_cache(wb, source_name: str):
    if not hasattr(wb, "_sheet_image_cache"):
        wb._sheet_image_cache = {}
    cache = wb._sheet_image_cache.get(source_name)
    if cache is None:
        cache = _build_image_cache(wb[source_name])
        wb._sheet_image_cache[source_name] = cache
    return cache


def _copy_sheet_with_images(wb, source_name: str, new_name: str):
    ws_src = wb[source_name]
    img_cache = _get_sheet_image_cache(wb, source_name)
    ws_new = wb.copy_worksheet(ws_src)
    ws_new.title = new_name

    try:
        ws_new._images = []
    except Exception:
        pass

    if img_cache:
        _apply_image_cache(ws_new, img_cache)
    return ws_new

def _clear_table_values_only(ws, row_start: int, row_end: int, cols=(1,3,5,7,8,9)):
    for r in range(row_start, row_end+1):
        for c in cols:
            cell = ws.cell(r,c)
            tl = _top_left_of_merged(ws, cell)
            if isinstance(tl, MergedCell):
                continue
            tl.value = None


def is_dongjin_agency(name: str) -> bool:
    a = re.sub(r"\s+", "", (name or "").strip())
    return "동진" in a


def make_filename(vessel: str, stays: List[CrewStay]) -> str:
    first_dt = min((s.checkin for s in stays), default=datetime.now())
    safe_vessel = "".join(ch for ch in str(vessel or "").strip() if ch.isalnum() or ch in ("-", "_")).strip() or "VESSEL"
    return f"INVOICE_{safe_vessel}_{first_dt.strftime('%Y%m%d')}.xlsx"


DONGJIN_MEAL_UNIT = 25000


def _dongjin_sort_room_key(room_no: str):
    s = str(room_no)
    m = re.match(r"^(\d+)(.*)$", s)
    if m:
        return (0, int(m.group(1)), m.group(2))
    return (1, s)


def _dongjin_ordered_rooms(room_groups: Dict[str, List[CrewStay]], room_payers: Dict[str, CrewStay]) -> List[str]:
    singles = []
    twins = []
    for rn in room_groups.keys():
        payer = room_payers.get(rn)
        if payer is not None and is_twin(getattr(payer, "room_type", "")):
            twins.append(rn)
        else:
            singles.append(rn)
    return sorted(singles, key=_dongjin_sort_room_key) + sorted(twins, key=_dongjin_sort_room_key)


def _dongjin_representative_room(room_groups: Dict[str, List[CrewStay]], room_payers: Dict[str, CrewStay]) -> Optional[str]:
    ordered = _dongjin_ordered_rooms(room_groups, room_payers)
    return ordered[0] if ordered else None


def _dongjin_parse_meal_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _dongjin_person_text(n: int) -> str:
    return f"{int(n)} Person" if int(n) == 1 else f"{int(n)} Person"


def _dongjin_guest_display(stays: List[CrewStay]) -> str:
    names = [str(s.crew_name or "").strip() for s in stays if str(s.crew_name or "").strip()]
    if not names:
        return ""
    if len(names) == 1:
        return names[0]
    return f"{names[0]} 외{len(names)-1}명"


def _dongjin_first_nationality(stays: List[CrewStay]) -> str:
    for s in stays:
        nat = str(getattr(s, "nationality", "") or "").strip()
        if nat:
            return nat
    return ""


def _dongjin_set_row_style_from(ws, src_row: int, dst_row: int, start_col: int = 1, end_col: int = 9):
    try:
        src_dim = ws.row_dimensions[src_row]
        dst_dim = ws.row_dimensions[dst_row]
        if src_dim.height:
            dst_dim.height = src_dim.height
    except Exception:
        pass

    for col in range(start_col, end_col + 1):
        src_cell = _top_left_of_merged(ws, ws.cell(src_row, col))
        dst_cell = _top_left_of_merged(ws, ws.cell(dst_row, col))
        if isinstance(src_cell, MergedCell) or isinstance(dst_cell, MergedCell):
            continue
        try:
            dst_cell._style = copy(src_cell._style)
        except Exception:
            pass
        try:
            dst_cell.number_format = src_cell.number_format
        except Exception:
            pass
        try:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.protection = copy(src_cell.protection)
        except Exception:
            pass


def _dongjin_clone_row_with_merges(ws, src_row: int, dst_row: int, max_col: int = 12):
    if src_row == dst_row:
        return

    merge_specs = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == src_row and mr.max_row == src_row:
            merge_specs.append((mr.min_col, mr.max_col))

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == dst_row and mr.max_row == dst_row:
            try:
                ws.unmerge_cells(str(mr))
            except Exception:
                pass

    try:
        src_dim = ws.row_dimensions[src_row]
        dst_dim = ws.row_dimensions[dst_row]
        if src_dim.height:
            dst_dim.height = src_dim.height
    except Exception:
        pass

    for col in range(1, max_col + 1):
        s = ws.cell(src_row, col)
        d = ws.cell(dst_row, col)
        try:
            d._style = copy(s._style)
            d.font = copy(s.font)
            d.fill = copy(s.fill)
            d.border = copy(s.border)
            d.number_format = s.number_format
            d.protection = copy(s.protection)
            d.alignment = copy(s.alignment)
        except Exception:
            pass
        if d.value is not None:
            d.value = None

    for start_col, end_col in merge_specs:
        try:
            ws.merge_cells(start_row=dst_row, start_column=start_col, end_row=dst_row, end_column=end_col)
        except Exception:
            pass


def _dongjin_prepare_meal_rows(ws, needed_rows: int):
    detail_start = 15
    footer_start = 25
    template_capacity = footer_start - detail_start
    extra = max(0, needed_rows - template_capacity)

    if extra > 0:
        ws.insert_rows(footer_start, amount=extra)

    last_detail_row = detail_start + max(needed_rows, 1) - 1

    # detail row template: copy merge/border/style from row 15 into every used detail row
    for r in range(detail_start, last_detail_row + 1):
        if r != detail_start:
            _dongjin_clone_row_with_merges(ws, detail_start, r, 12)
        else:
            _dongjin_clone_row_with_merges(ws, detail_start, detail_start, 12)
        _dongjin_set_row_style_from(ws, detail_start, r, 1, 9)

    # summary/footer rows must keep the same visual structure after inserted rows
    lunch_row = 28 + extra
    dinner_row = 29 + extra
    template_lunch_row = 28
    template_dinner_row = 29
    if lunch_row != template_lunch_row:
        _dongjin_clone_row_with_merges(ws, template_lunch_row, lunch_row, 12)
        _dongjin_set_row_style_from(ws, template_lunch_row, lunch_row, 1, 9)
    if dinner_row != template_dinner_row:
        _dongjin_clone_row_with_merges(ws, template_dinner_row, dinner_row, 12)
        _dongjin_set_row_style_from(ws, template_dinner_row, dinner_row, 1, 9)

    return extra, last_detail_row


def _dongjin_find_summary_footer_row(ws, detail_start: int = 14) -> int:
    footer_keywords = (
        "Our Bank Account",
        "BANK NAME",
        "ACCOUNT NO",
        "I agree that",
        "본인은",
        "지정회사",
        "지정단체",
    )
    for r in range(detail_start + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in (1, 3, 5, 9)]
        joined = " ".join(str(v) for v in vals if v not in (None, ""))
        if any(key in joined for key in footer_keywords):
            return r
    return 30


def _dongjin_find_total_formula_row(ws, default_row: int) -> int:
    for r in range(max(20, default_row - 3), min(ws.max_row, default_row + 8) + 1):
        v = ws.cell(r, 9).value
        if isinstance(v, str) and "SUM(" in v.upper():
            return r
    for r in range(20, min(ws.max_row, 60) + 1):
        v = ws.cell(r, 9).value
        if isinstance(v, str) and "SUM(" in v.upper():
            return r
    return default_row


def _dongjin_prepare_summary_rows(ws, needed_rows: int, detail_start: int = 14, max_col: int = 9):
    footer_start = _dongjin_find_summary_footer_row(ws, detail_start)
    if footer_start <= detail_start:
        footer_start = detail_start + 16

    template_capacity = max(1, footer_start - detail_start)
    extra = max(0, needed_rows - template_capacity)

    if extra > 0:
        ws.insert_rows(footer_start, amount=extra)

    last_detail_row = detail_start + max(needed_rows, 1) - 1

    for r in range(detail_start + 1, last_detail_row + 1):
        _dongjin_clone_row_with_merges(ws, detail_start, r, max_col)
        _dongjin_set_row_style_from(ws, detail_start, r, 1, max_col)

    clear_end = max(last_detail_row, footer_start + extra - 1)
    return extra, last_detail_row, clear_end

def _dongjin_template_sources(wb):
    names = list(wb.sheetnames)

    def first_existing(cands):
        for n in cands:
            if n in names:
                return n
        return None

    hotel_summary = first_existing(["HOTEL", "BILL"])
    hotel_detail = first_existing(["HOTEL BILL", "ROOM"])

    restaurant_candidates = [n for n in ("RESTAURANT", "RESTAURANT BILL") if n in names]
    meal_summary = None
    meal_detail = None

    for n in restaurant_candidates:
        ws = wb[n]
        if ws.max_column <= 9 and meal_summary is None:
            meal_summary = n
        if ws.max_column >= 12 and meal_detail is None:
            meal_detail = n

    if meal_summary is None:
        meal_summary = first_existing(["RESTAURANT", "RESTAURANT BILL"])
    if meal_detail is None:
        for n in reversed(restaurant_candidates):
            if n != meal_summary:
                meal_detail = n
                break

    missing = [label for label, val in [
        ("hotel_summary", hotel_summary),
        ("hotel_detail", hotel_detail),
        ("meal_summary", meal_summary),
        ("meal_detail", meal_detail),
    ] if not val]
    if missing:
        raise ValueError(
            "동진 템플릿 시트 구성을 찾지 못했습니다. "
            f"현재 시트: {wb.sheetnames}"
        )

    return {
        "hotel_summary": hotel_summary,
        "hotel_detail": hotel_detail,
        "meal_summary": meal_summary,
        "meal_detail": meal_detail,
    }


def _dongjin_fill_headers(ws, agency_en: str, vessel: str, room_no=None, checkin=None, checkout=None):
    safe_set_addr(ws, "C7", agency_en)
    safe_set_addr(ws, "C8", vessel)
    if room_no is not None:
        safe_set_addr(ws, "I8", room_no)
    if checkin is not None:
        safe_set_addr(ws, "C10", checkin.date())
        safe_set_addr(ws, "C11", checkin.strftime("%H:%M"))
    if checkout is not None:
        safe_set_addr(ws, "I10", checkout.date())
        safe_set_addr(ws, "I11", checkout.strftime("%H:%M"))


def _dongjin_room_names(room_stays: List[CrewStay]) -> str:
    names = [str(s.crew_name or "").strip() for s in room_stays if str(s.crew_name or "").strip()]
    return "\n".join(names)


def _dongjin_room_nationalities(room_stays: List[CrewStay]) -> str:
    vals = [str(getattr(s, "nationality", "") or "").strip() for s in room_stays if str(getattr(s, "nationality", "") or "").strip()]
    return "\n".join(vals)


def _dj_top_left(ws, row: int, col: int):
    return _top_left_of_merged(ws, ws.cell(row, col))


def _dj_set_center(ws, row: int, col: int):
    cell = _dj_top_left(ws, row, col)
    if isinstance(cell, MergedCell):
        return
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _dj_set_left_wrap(ws, row: int, col: int):
    cell = _dj_top_left(ws, row, col)
    if isinstance(cell, MergedCell):
        return
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _dj_apply_date_fmt(ws, row: int, col: int):
    cell = _dj_top_left(ws, row, col)
    if isinstance(cell, MergedCell):
        return
    cell.number_format = EXCEL_DATE_FMT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _dj_apply_amount_fmt(ws, row: int, col: int, center: bool = True):
    cell = _dj_top_left(ws, row, col)
    if isinstance(cell, MergedCell):
        return
    cell.number_format = EXCEL_AMOUNT_ZERO_DASH_FMT
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _dj_style_meal_detail_row(ws, row: int):
    _dj_set_center(ws, row, 1)
    _dj_apply_date_fmt(ws, row, 3)
    for col in (5, 7, 8, 9):
        _dj_apply_amount_fmt(ws, row, col, center=True)


def _dj_style_summary_row(ws, row: int):
    _dj_set_center(ws, row, 1)
    _dj_apply_date_fmt(ws, row, 3)
    _dj_set_left_wrap(ws, row, 5)
    _dj_apply_amount_fmt(ws, row, 9, center=False)



def _dongjin_fill_hotel_detail(ws, room_no: str, room_stays: List[CrewStay], payer: CrewStay, payer_rate: RateRow, agency_en: str):
    _dongjin_fill_headers(
        ws, agency_en, payer.vessel,
        room_no=int(room_no) if str(room_no).isdigit() else room_no,
        checkin=payer.checkin, checkout=payer.checkout
    )
    guest_names = _dongjin_room_names(room_stays) or (payer.crew_name or "")
    guest_nations = _dongjin_room_nationalities(room_stays) or (getattr(payer, "nationality", "") or "")

    safe_set_addr(ws, "C9", guest_names)
    safe_set_wrap_left(ws, "C9")
    safe_set_addr(ws, "I9", guest_nations)
    safe_set_wrap_left(ws, "I9")

    try:
        name_lines = max(1, str(guest_names).count("\n") + 1) if str(guest_names).strip() else 1
        nation_lines = max(1, str(guest_nations).count("\n") + 1) if str(guest_nations).strip() else 1
        needed_lines = max(name_lines, nation_lines)
        ws.row_dimensions[9].height = max(ws.row_dimensions[9].height or 18, 18 * needed_lines)
    except Exception:
        pass

    _clear_table_values_only(ws, 14, 38, cols=(1, 3, 5, 9))

    base_nights, extra_out, has_late, extra_in, has_early = calc_room_parts(payer.checkin, payer.checkout)
    row = 14
    no = 1
    acc_desc = "Accommodation Charge(Breakfast included)"

    def add_line(dd, desc, amt):
        nonlocal row, no
        safe_set_cell_merged(ws, row, 1, no)
        safe_set_cell_merged(ws, row, 3, dd)
        safe_set_cell_merged(ws, row, 5, desc)
        safe_set_cell_merged(ws, row, 9, amt)
        try:
            _dj_set_center(ws, row, 1)
            _dj_apply_date_fmt(ws, row, 3)
            _dj_apply_amount_fmt(ws, row, 9, center=False)
        except Exception:
            pass
        row += 1
        no += 1

    if extra_in:
        add_line(payer.checkin.date(), acc_desc, payer_rate.nightly)

    for i in range(max(base_nights, 1)):
        add_line(payer.checkin.date() + timedelta(days=i), acc_desc, payer_rate.nightly)

    if extra_out:
        add_line(payer.checkout.date(), acc_desc, payer_rate.nightly)

    if has_early:
        add_line(payer.checkin.date(), "Accommodation Charge(Early check in)", payer_rate.early_0801_0959)

    if has_late:
        add_line(payer.checkout.date(), "Accommodation Charge(Late check out)", payer_rate.late_1201_1400)


def _dongjin_collect_meal_rows(room_stays: List[CrewStay]):
    per_date: Dict[date, Dict[str, int]] = {}
    lunch_people = set()
    dinner_people = set()
    lunch_max = 0
    dinner_max = 0

    for s in room_stays:
        meals = getattr(s, "meals_by_date", None) or {}
        crew = str(s.crew_name or "").strip()
        for d_key, flags in meals.items():
            dd = _dongjin_parse_meal_date(d_key)
            if dd is None:
                continue
            try:
                _, l, d = flags
            except Exception:
                continue

            cur = per_date.setdefault(dd, {"lunch_cnt": 0, "dinner_cnt": 0})
            if l:
                cur["lunch_cnt"] += 1
                if crew:
                    lunch_people.add(crew)
            if d:
                cur["dinner_cnt"] += 1
                if crew:
                    dinner_people.add(crew)

    for vals in per_date.values():
        lunch_max = max(lunch_max, vals.get("lunch_cnt", 0))
        dinner_max = max(dinner_max, vals.get("dinner_cnt", 0))
        vals["breakfast_amt"] = 0
        vals["lunch_amt"] = vals.get("lunch_cnt", 0) * DONGJIN_MEAL_UNIT
        vals["dinner_amt"] = vals.get("dinner_cnt", 0) * DONGJIN_MEAL_UNIT
        vals["total_amt"] = vals["breakfast_amt"] + vals["lunch_amt"] + vals["dinner_amt"]

    return per_date, lunch_max, dinner_max, len(lunch_people), len(dinner_people)



def _dongjin_fill_meal_detail(ws, room_no: str, meal_stays: List[CrewStay], payer: CrewStay, agency_en: str):
    all_checkins = [s.checkin for s in meal_stays] or [payer.checkin]
    all_checkouts = [s.checkout for s in meal_stays] or [payer.checkout]
    start_dt = min(all_checkins)
    end_dt = max(all_checkouts)

    _dongjin_fill_headers(
        ws, agency_en, payer.vessel,
        room_no=int(room_no) if str(room_no).isdigit() else room_no,
        checkin=start_dt, checkout=end_dt
    )
    meal_guest = _dongjin_guest_display(meal_stays) or (payer.crew_name or "")
    meal_nat = _dongjin_first_nationality(meal_stays) or (getattr(payer, "nationality", "") or "")

    safe_set_addr(ws, "C9", meal_guest)
    safe_set_wrap_left(ws, "C9")
    safe_set_addr(ws, "I9", meal_nat)
    safe_set_wrap_left(ws, "I9")

    try:
        name_lines = max(1, str(meal_guest).count("\n") + 1) if str(meal_guest).strip() else 1
        nat_lines = max(1, str(meal_nat).count("\n") + 1) if str(meal_nat).strip() else 1
        needed_lines = max(name_lines, nat_lines)
        ws.row_dimensions[9].height = max(ws.row_dimensions[9].height or 18, 18 * needed_lines)
    except Exception:
        pass

    per_date, lunch_max, dinner_max, lunch_people, dinner_people = _dongjin_collect_meal_rows(meal_stays)

    d0 = start_dt.date()
    d1 = end_dt.date()
    total_days = (d1 - d0).days
    if total_days < 0:
        total_days = 0
    needed_rows = total_days + 1

    extra_rows, last_detail_row = _dongjin_prepare_meal_rows(ws, needed_rows)
    _clear_table_values_only(ws, 15, max(last_detail_row, 38 + extra_rows), cols=(1, 3, 5, 7, 8, 9))
    safe_set_cell_merged(ws, 14, 5, "Breakfast")
    safe_set_cell_merged(ws, 14, 7, "Lunch")
    safe_set_cell_merged(ws, 14, 8, "Dinner")

    row = 15
    no = 1
    for i in range(total_days + 1):
        dd = d0 + timedelta(days=i)
        vals = per_date.get(dd, {})
        b_amt = 0
        l_amt = int(vals.get("lunch_amt", 0))
        d_amt = int(vals.get("dinner_amt", 0))

        safe_set_cell_merged(ws, row, 1, no)
        safe_set_cell_merged(ws, row, 3, dd)
        safe_set_cell_merged(ws, row, 5, b_amt)
        safe_set_cell_merged(ws, row, 7, l_amt)
        safe_set_cell_merged(ws, row, 8, d_amt)
        safe_set_cell_merged(ws, row, 9, f"=SUM(E{row},G{row}:H{row})")
        try:
            _dj_style_meal_detail_row(ws, row)
        except Exception:
            pass
        row += 1
        no += 1

    lunch_row = 28 + extra_rows
    dinner_row = 29 + extra_rows
    try:
        safe_set_addr(ws, f"G{lunch_row}", "LUNCH")
        safe_set_addr(ws, f"G{dinner_row}", "DINNER")
        safe_set_addr(ws, f"H{lunch_row}", f"{lunch_max}Person" if lunch_max else "0Person")
        safe_set_addr(ws, f"H{dinner_row}", f"{dinner_max}Person" if dinner_max else "0Person")
        _dj_set_center(ws, lunch_row, 7)
        _dj_set_center(ws, lunch_row, 8)
        _dj_set_center(ws, dinner_row, 7)
        _dj_set_center(ws, dinner_row, 8)
    except Exception:
        pass


def _dongjin_hotel_amount(payer: CrewStay, payer_rate: RateRow) -> int:
    room_total, late_fee, early_fee = calc_room_charge(payer.checkin, payer.checkout, payer_rate)
    return int(room_total + late_fee + early_fee)



def _dongjin_fill_hotel_summary(ws, room_groups: Dict[str, List[CrewStay]], room_payers: Dict[str, CrewStay], rates: Dict[Tuple[str, str], RateRow], agency_en: str, vessel: str):
    if not room_groups:
        return
    ordered_rooms = _dongjin_ordered_rooms(room_groups, room_payers)
    all_stays = [s for rn in ordered_rooms for s in room_groups[rn]]
    min_checkin = min(s.checkin for s in all_stays)
    max_checkout = max(s.checkout for s in all_stays)

    single_count = sum(1 for rn in ordered_rooms if not is_twin(getattr(room_payers[rn], "room_type", "")))
    twin_count = len(ordered_rooms) - single_count

    safe_set_addr(ws, "C7", agency_en)
    safe_set_addr(ws, "C8", vessel)
    if single_count and twin_count:
        safe_set_addr(ws, "I8", f"{single_count} Room - Single")
        safe_set_addr(ws, "I9", f"{twin_count} Room - Twin")
    elif twin_count and not single_count:
        safe_set_addr(ws, "I8", f"{twin_count} Room - Twin")
        safe_set_addr(ws, "I9", None)
    else:
        safe_set_addr(ws, "I8", f"{single_count} Room")
        safe_set_addr(ws, "I9", None)
    safe_set_addr(ws, "C10", min_checkin.date())
    safe_set_addr(ws, "C11", min_checkin.strftime("%H:%M"))
    safe_set_addr(ws, "I10", max_checkout.date())
    safe_set_addr(ws, "I11", max_checkout.strftime("%H:%M"))

    _, last_detail_row, clear_end = _dongjin_prepare_summary_rows(ws, len(ordered_rooms), detail_start=14, max_col=9)
    _clear_table_values_only(ws, 14, max(34, clear_end), cols=(1, 3, 5, 9))

    row = 14
    no = 1
    for rn in ordered_rooms:
        payer = room_payers[rn]
        payer_rate = rates[(payer.agency_name, payer.room_type)]
        amt = _dongjin_hotel_amount(payer, payer_rate)
        guest = _dongjin_room_names(room_groups[rn]) or str(payer.crew_name or "").strip() or f"ROOM {rn}"
        safe_set_cell_merged(ws, row, 1, no)
        safe_set_cell_merged(ws, row, 3, payer.checkout.date())
        safe_set_cell_merged(ws, row, 5, guest)
        safe_set_cell_merged(ws, row, 9, amt)
        try:
            _dj_style_summary_row(ws, row)
            lines = max(1, str(guest).count("\n") + 1)
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 18, 18 * lines)
        except Exception:
            pass
        row += 1
        no += 1

    total_row = _dongjin_find_total_formula_row(ws, 35)
    try:
        safe_set_addr(ws, f"I{total_row}", f"=SUM(I14:I{last_detail_row})")
        _dj_apply_amount_fmt(ws, total_row, 9, center=False)
    except Exception:
        pass


def _dongjin_room_meal_total(room_stays: List[CrewStay]) -> int:
    per_date, _, _, _, _ = _dongjin_collect_meal_rows(room_stays)
    return sum(int(v.get("total_amt", 0)) for v in per_date.values())



def _dongjin_fill_meal_summary(
    ws,
    room_groups: Dict[str, List[CrewStay]],
    room_payers: Dict[str, CrewStay],
    agency_en: str,
    vessel: str,
    rep_room_no: Optional[str] = None,
):
    if not room_groups:
        return
    ordered_rooms = _dongjin_ordered_rooms(room_groups, room_payers)
    if rep_room_no is None:
        rep_room_no = ordered_rooms[0] if ordered_rooms else None
    all_stays = [s for rn in ordered_rooms for s in room_groups[rn]]
    min_checkin = min(s.checkin for s in all_stays)
    max_checkout = max(s.checkout for s in all_stays)

    safe_set_addr(ws, "C7", agency_en)
    safe_set_addr(ws, "C8", vessel)
    safe_set_addr(ws, "I8", _dongjin_person_text(len(all_stays)))
    safe_set_addr(ws, "C10", min_checkin.date())
    safe_set_addr(ws, "C11", min_checkin.strftime("%H:%M"))
    safe_set_addr(ws, "I10", max_checkout.date())
    safe_set_addr(ws, "I11", max_checkout.strftime("%H:%M"))

    grand_meal_total = _dongjin_room_meal_total(all_stays)

    _, last_detail_row, clear_end = _dongjin_prepare_summary_rows(ws, len(ordered_rooms), detail_start=14, max_col=9)
    _clear_table_values_only(ws, 14, max(34, clear_end), cols=(1, 3, 5, 9))

    row = 14
    no = 1
    for rn in ordered_rooms:
        payer = room_payers[rn]
        amt = grand_meal_total if rn == rep_room_no else 0
        guest = _dongjin_room_names(room_groups[rn]) or str(payer.crew_name or "").strip() or f"ROOM {rn}"
        safe_set_cell_merged(ws, row, 1, no)
        safe_set_cell_merged(ws, row, 3, payer.checkout.date())
        safe_set_cell_merged(ws, row, 5, guest)
        safe_set_cell_merged(ws, row, 9, amt)
        try:
            _dj_style_summary_row(ws, row)
            lines = max(1, str(guest).count("\n") + 1)
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 18, 18 * lines)
        except Exception:
            pass
        row += 1
        no += 1

    total_row = _dongjin_find_total_formula_row(ws, 36)
    try:
        safe_set_addr(ws, f"I{total_row}", f"=SUM(I14:I{last_detail_row})")
        _dj_apply_amount_fmt(ws, total_row, 9, center=False)
    except Exception:
        pass



def _detect_print_last_row(ws, minimum_row: int = 46) -> int:
    last_row = 1

    try:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    if cell.row > last_row:
                        last_row = cell.row
    except Exception:
        pass

    try:
        for rng in ws.merged_cells.ranges:
            if rng.max_row > last_row:
                last_row = rng.max_row
    except Exception:
        pass

    try:
        for img in getattr(ws, "_images", []):
            anc = getattr(img, "anchor", None)
            if hasattr(anc, "_from"):
                last_row = max(last_row, anc._from.row + 1)
            if hasattr(anc, "_to"):
                last_row = max(last_row, anc._to.row + 1)
    except Exception:
        pass

    return max(last_row, minimum_row)


def _apply_one_page_print_settings(ws, minimum_row: int = 46) -> None:
    last_row = _detect_print_last_row(ws, minimum_row=minimum_row)

    try:
        ws.print_area = f"A1:I{last_row}"
    except Exception:
        pass

    try:
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties()
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    except Exception:
        pass

    try:
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.scale = None
    except Exception:
        pass

    try:
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        ws.page_margins.header = 0.1
        ws.page_margins.footer = 0.1
    except Exception:
        pass

    try:
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        ws.print_options.gridLines = False
    except Exception:
        pass

    try:
        ws.row_breaks.brk = []
        ws.col_breaks.brk = []
    except Exception:
        pass


def _apply_one_page_print_settings_to_workbook(wb) -> None:
    for ws in wb.worksheets:
        minimum_row = 46
        try:
            title = (ws.title or "").strip().upper()
        except Exception:
            title = ""

        if title in {"HOTEL", "RESTAURANT", "BILL"}:
            minimum_row = 43
        elif title.startswith("BILL-") or title.startswith("MEAL-"):
            minimum_row = 46

        _apply_one_page_print_settings(ws, minimum_row=minimum_row)


def _dongjin_save_invoice(vessel: str, stays: List[CrewStay], rates: Dict[Tuple[str, str], RateRow]) -> Path:
    tpl_dj = resolve_template_path("동진")
    if not tpl_dj.exists():
        raise FileNotFoundError(f"동진 템플릿 파일이 없습니다: {tpl_dj}")

    wb = load_workbook(tpl_dj)
    rehydrate_images(wb)

    src = _dongjin_template_sources(wb)

    room_groups: Dict[str, List[CrewStay]] = defaultdict(list)
    for s in stays:
        room_groups[str(s.room_no)].append(s)

    room_payers: Dict[str, CrewStay] = {}
    for rn, lst in room_groups.items():
        payer = next((x for x in lst if getattr(x, "charge_room", False)), lst[0])
        room_payers[rn] = payer

    ordered_rooms = _dongjin_ordered_rooms(room_groups, room_payers)
    rep_room_no = _dongjin_representative_room(room_groups, room_payers)

    agencies_kor = sorted({(s.agency_name or "").strip() for s in stays})
    agency_disp_kor = agencies_kor[0] if len(agencies_kor) == 1 else "MIXED"
    agency_en = agency_display_eng_for_group(stays, rates, agency_disp_kor)

    hotel_detail_source = src["hotel_detail"]
    meal_detail_source = src["meal_detail"]

    for rn in ordered_rooms:
        payer = room_payers[rn]
        room_stays = room_groups[rn]
        payer_rate = rates[(payer.agency_name, payer.room_type)]
        ws_h = _copy_sheet_with_images(wb, hotel_detail_source, f"BILL-{rn}")
        _dongjin_fill_hotel_detail(ws_h, rn, room_stays, payer, payer_rate, agency_en)

    if rep_room_no:
        rep_payer = room_payers[rep_room_no]
        ws_m = _copy_sheet_with_images(wb, meal_detail_source, f"MEAL-{rep_room_no}")
        _dongjin_fill_meal_detail(ws_m, rep_room_no, stays, rep_payer, agency_en)

    for n in [hotel_detail_source, meal_detail_source]:
        if n in wb.sheetnames:
            try:
                wb.remove(wb[n])
            except Exception:
                pass

    hotel_summary_source = src["hotel_summary"]
    meal_summary_source = src["meal_summary"]

    if hotel_summary_source in wb.sheetnames:
        wb[hotel_summary_source].title = "HOTEL"
    if meal_summary_source in wb.sheetnames and wb[meal_summary_source].title != "RESTAURANT":
        wb[meal_summary_source].title = "RESTAURANT"

    _dongjin_fill_hotel_summary(wb["HOTEL"], room_groups, room_payers, rates, agency_en, vessel)
    _dongjin_fill_meal_summary(wb["RESTAURANT"], room_groups, room_payers, agency_en, vessel, rep_room_no=rep_room_no)

    rehydrate_images(wb)
    _apply_one_page_print_settings_to_workbook(wb)

    out_dir = get_output_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = make_filename(vessel, stays)
    out_path = out_dir / filename
    attempt = 0
    while True:
        try:
            wb.save(out_path)
            return out_path
        except PermissionError:
            attempt += 1
            out_path = out_dir / f"{Path(filename).stem}_{attempt}.xlsx"
            if attempt >= 50:
                raise


def load_rates(path: Path) -> Dict[Tuple[str, str], RateRow]:
    """
    rates.xlsx / 요율 시트 구조:
      1행: 영문키 헤더
      2행: 한글 설명(있어도 됨) -> 스킵
      3행~: 데이터
    필요 헤더:
      agency_name | agency_eng | room_type | nightly | late_1201_1400 | early_0801_0959 | breakfast | lunch | dinner
    """
    if not path.exists():
        raise FileNotFoundError(f"요율 파일이 없습니다: {path}")

    wb = load_workbook(path, data_only=True)
    if RATES_SHEET not in wb.sheetnames:
        raise ValueError(f"'{RATES_SHEET}' 시트가 없습니다. 현재 시트: {wb.sheetnames}")

    ws = wb[RATES_SHEET]

    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip()] = col

    required = [
        "agency_name", "agency_eng", "room_type",
        "nightly", "late_1201_1400", "early_0801_0959",
        "breakfast", "lunch", "dinner",
    ]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(
            "요율 시트 헤더가 부족합니다.\n"
            f"누락: {missing}\n\n"
            "필요 헤더 예시:\n"
            "agency_name | agency_eng | room_type | nightly | late_1201_1400 | early_0801_0959 | breakfast | lunch | dinner"
        )

    def _cell(r: int, name: str):
        return ws.cell(row=r, column=headers[name]).value

    rates: Dict[Tuple[str, str], RateRow] = {}

    for r in range(3, ws.max_row + 1):  # ✅ 3행부터 데이터
        agency = _cell(r, "agency_name")
        room = _cell(r, "room_type")
        if not agency or not room:
            continue

        agency_kor = str(agency).strip()
        room_type = str(room).strip()

        agency_eng = _cell(r, "agency_eng")
        agency_eng = str(agency_eng).strip() if agency_eng else ""

        rates[(agency_kor, room_type)] = RateRow(
            agency_eng=agency_eng,
            nightly=_to_int(_cell(r, "nightly")),
            late_1201_1400=_to_int(_cell(r, "late_1201_1400")),
            early_0801_0959=_to_int(_cell(r, "early_0801_0959")),
            breakfast=_to_int(_cell(r, "breakfast")),
            lunch=_to_int(_cell(r, "lunch")),
            dinner=_to_int(_cell(r, "dinner")),
        )

    if not rates:
        raise ValueError("요율 시트에서 유효한 요율 행을 찾지 못했습니다. (3행부터 데이터가 있는지 확인)")
    return rates


def available_agencies(rates: Dict[Tuple[str, str], RateRow]) -> List[str]:
    return sorted({k[0] for k in rates.keys()})


def available_room_types_for_agency(rates: Dict[Tuple[str, str], RateRow], agency: str) -> List[str]:
    return sorted({k[1] for k in rates.keys() if k[0] == agency})


def agency_display_eng_for_group(stays: List[CrewStay], rates: Dict[Tuple[str, str], RateRow], agency_kor: str) -> str:
    """BILL 등에 찍을 대리점 표기(영문)."""
    if agency_kor == "MIXED":
        return "MIXED"
    for s in stays:
        if s.agency_name == agency_kor:
            rr = rates.get((s.agency_name, s.room_type))
            if rr and rr.agency_eng.strip():
                return rr.agency_eng.strip()
            break
    return agency_kor


# =========================
# Room charge logic (EARLY/LATE)
# =========================
def calc_room_parts(checkin: datetime, checkout: datetime) -> Tuple[int, int, bool, int, bool]:
    """
    Returns:
      base_nights, extra_checkout_nights, has_late_fee, extra_checkin_nights, has_early_fee
    """
    days = (checkout.date() - checkin.date()).days
    base_nights = 1 if days <= 0 else days

    out_t = checkout.time()
    has_late = (time(12, 1) <= out_t <= time(14, 0))
    extra_out = 1 if (out_t >= time(14, 1)) else 0

    in_t = checkin.time()
    has_early = (time(8, 1) <= in_t <= time(9, 59))
    extra_in = 1 if (in_t <= time(8, 0)) else 0

    return base_nights, extra_out, has_late, extra_in, has_early


def calc_room_charge(checkin: datetime, checkout: datetime, rate: RateRow) -> Tuple[int, int, int]:
    base_nights, extra_out, has_late, extra_in, has_early = calc_room_parts(checkin, checkout)
    room_total = (base_nights + extra_in + extra_out) * rate.nightly
    late_fee = rate.late_1201_1400 if has_late else 0
    early_fee = rate.early_0801_0959 if has_early else 0
    return room_total, late_fee, early_fee


def meal_amounts(stays: List[CrewStay], rates: Dict[Tuple[str, str], RateRow]) -> Tuple[int, int, int]:
    b_amt = l_amt = d_amt = 0
    for s in stays:
        rate = rates[(s.agency_name, s.room_type)]
        for _, (b, l, d) in s.meals_by_date.items():
            if b:
                b_amt += rate.breakfast
            if l:
                l_amt += rate.lunch
            if d:
                d_amt += rate.dinner
    return b_amt, l_amt, d_amt


def meal_amounts_by_date(stays: List[CrewStay], rates: Dict[Tuple[str, str], RateRow], include_breakfast: bool = True) -> Dict[date, Tuple[int, int, int]]:
    """Returns {date: (breakfast_amount, lunch_amount, dinner_amount)} summed across crew."""
    out: Dict[date, List[int]] = {}
    for s in stays:
        rate = rates[(s.agency_name, s.room_type)]
        for d_str, (b, l, d) in s.meals_by_date.items():
            try:
                dd = datetime.strptime(d_str, "%Y-%m-%d").date()
            except Exception:
                continue
            if dd not in out:
                out[dd] = [0, 0, 0]
            if include_breakfast and b:
                out[dd][0] += rate.breakfast
            if l:
                out[dd][1] += rate.lunch
            if d:
                out[dd][2] += rate.dinner
    return {k: (v[0], v[1], v[2]) for k, v in out.items()}


# =========================
# Merged-cell safe setters
# =========================
def _top_left_of_merged(ws, cell):
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell


def safe_set_cell_merged(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    tl = _top_left_of_merged(ws, cell)
    if isinstance(tl, MergedCell):
        return
    tl.value = value

def safe_set_addr(ws, addr: str, value):
    c = ws[addr]
    c2 = _top_left_of_merged(ws, c)
    c2.value = value


def safe_set_rc(ws, row: int, col: int, value):
    c = ws.cell(row=row, column=col)
    if isinstance(c, MergedCell):
        c = _top_left_of_merged(ws, c)
    c.value = value


def safe_set_wrap(ws, addr: str):
    """✅ 줄바꿈 표시(wrap) 강제 적용"""
    c = ws[addr]
    c2 = _top_left_of_merged(ws, c)
    c2.alignment = Alignment(wrap_text=True, vertical="top")

def safe_set_wrap_left(ws, addr: str):
    """줄바꿈 + 왼쪽 정렬"""
    c = ws[addr]
    c2 = _top_left_of_merged(ws, c)
    c2.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")



# =========================
# Image handling (logo 유지)
# =========================

BASE_ROOM_IMAGE_CACHE: List[Tuple[bytes, Optional[int], Optional[int], object]] = []


def _build_image_cache(src_ws):
    cache = []
    for img in list(getattr(src_ws, "_images", [])):
        try:
            raw = img._data()
        except Exception:
            continue

        w = getattr(img, "width", None)
        h = getattr(img, "height", None)

        anc = None
        if hasattr(img, "anchor"):
            try:
                anc = copy(img.anchor)
            except Exception:
                anc = None

        cache.append((raw, w, h, anc))
    return cache


def _apply_image_cache(dst_ws, cache):
    if not cache:
        return

    for raw, w, h, anc in cache:
        try:
            bio = io.BytesIO(raw)
            new_img = XLImage(bio)
            # keep stream alive after save
            if not hasattr(dst_ws.parent, "_img_streams"):
                dst_ws.parent._img_streams = []
            dst_ws.parent._img_streams.append(bio)
        except Exception:
            continue

        if w is not None:
            try:
                new_img.width = w
            except Exception:
                pass
        if h is not None:
            try:
                new_img.height = h
            except Exception:
                pass

        if anc is not None:
            try:
                new_img.anchor = copy(anc)
            except Exception:
                pass

        try:
            dst_ws.add_image(new_img)
        except Exception:
            pass


def rehydrate_images(wb):
    """이미지 스트림 유지(저장 시 사라지는 문제 완화)."""
    wb._img_streams = []

    for ws in wb.worksheets:
        imgs = list(getattr(ws, "_images", []))
        if not imgs:
            continue

        ws._images = []
        for img in imgs:
            try:
                raw = img._data()
            except Exception:
                continue

            bio = io.BytesIO(raw)
            new_img = XLImage(bio)
            wb._img_streams.append(bio)

            try:
                new_img.width = img.width
                new_img.height = img.height
            except Exception:
                pass

            if hasattr(img, "anchor"):
                try:
                    new_img.anchor = copy(img.anchor)
                except Exception:
                    pass

            try:
                ws.add_image(new_img)
            except Exception:
                pass

# =========================
# Excel util
# =========================
def _clear_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _set_total_amount_cells(ws, mode: str = "default") -> None:
    """TOTAL 영역에 값을 '수식'으로 넣기.
    - default(협성/동진 제외): BILL I32 =SUM(I14:I31)
    - hyupsung(협성): BILL I35 =SUM(I14:I34)
    ※ BALANCE DUE에는 넣지 않음.
    """
    try:
        if mode == "hyupsung":
            ws["I35"].value = "=SUM(I14:I34)"
        else:
            ws["I32"].value = "=SUM(I14:I31)"
    except Exception:
        pass
def _set_total_invoice_default_total(ws) -> None:
    """기본 템플릿 TOTAL INVOICE(HOTEL) 시트:
    - I14~I30의 AMOUNT 합계를 I32에만 기록
    - 다른 TOTAL/BALANCE 셀은 건드리지 않음
    """
    total_val = 0
    for r in range(14, 31):
        v = ws.cell(row=r, column=9).value  # I
        try:
            if v is None or v == "":
                continue
            total_val += int(v)
        except Exception:
            # 숫자 아닌 경우 무시
            continue
    try:
        cell = ws["I32"]
        tl = _top_left_of_merged(ws, cell)
        if not isinstance(tl, MergedCell):
            tl.value = total_val
            tl.number_format = "#,##0"
    except Exception:
        pass

def _clear_bill_table_cells(ws, start_row: int, end_row: int):
    """BILL 시트에서 '입력 테이블' 부분만 지우기.
    ✅ 템플릿에 있는 고정 문구/라벨/수식은 유지하고,
    우리가 쓰는 칸(A,C,E,I)만 비웁니다.

    협성(BILL) 템플릿은 하단에 고정 문구(예: 'I agree that ...')가 있으므로
    해당 문구가 시작되는 행을 자동으로 찾아 그 위까지만 정리합니다.
    """
    # footer 자동 탐지 (영문/한글 모두)
    footer_row = None
    for r in range(start_row, end_row + 1):
        v = ws.cell(row=r, column=1).value
        if not v:
            continue
        s = str(v)
        if ("I agree that" in s) or ("liability" in s) or ("본인은" in s) or ("지정" in s):
            footer_row = r
            break

    if footer_row is not None:
        end_row = min(end_row, footer_row - 1)

    cols = (1, 3, 5, 9)  # A, C, E, I
    for r in range(start_row, end_row + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None





def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 12):
    for c in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        d._style = copy(s._style)
        d.font = copy(s.font)
        d.border = copy(s.border)
        d.fill = copy(s.fill)
        d.number_format = s.number_format
        d.protection = copy(s.protection)
        d.alignment = copy(s.alignment)


def _ensure_room_sheet(wb, room_no: str, base_room_sheet_name: str) -> str:
    """
    룸 시트 생성/재사용.
    ✅ 로고 누락 방지: BASE_ROOM_IMAGE_CACHE 붙이기
    """
    if room_no in wb.sheetnames:
        ws = wb[room_no]
        if len(getattr(ws, "_images", [])) == 0 and BASE_ROOM_IMAGE_CACHE:
            _apply_image_cache(ws, BASE_ROOM_IMAGE_CACHE)
        return room_no

    if base_room_sheet_name not in wb.sheetnames:
        raise ValueError(f"템플릿에 '{base_room_sheet_name}' 시트가 없어 룸 시트를 자동 생성할 수 없습니다.")

    base = wb[base_room_sheet_name]
    new_ws = wb.copy_worksheet(base)
    new_ws.title = room_no

    try:
        new_ws._images = []
    except Exception:
        pass

    if BASE_ROOM_IMAGE_CACHE:
        _apply_image_cache(new_ws, BASE_ROOM_IMAGE_CACHE)

    return room_no


# =========================
# Room sheet fill (DEFAULT)
# =========================
def _fill_room_sheet_default(
    ws,
    room_no: str,
    room_type: str,
    agency_display_for_excel: str,
    vessel: str,
    stays_in_room: List[CrewStay],
    rates: Dict[Tuple[str, str], RateRow],
):
    """기본 대리점(협성/동진 제외) ROOM 시트 채우기.

    ✅ 참고 파일(신양-1명-TRAMINO INDEPENDENT-1-11~1-13-호텔.xlsx) 구조에 맞춤:
      - C10(Arrival Date), C11(Arrival Time), I10(Departure Date), I11(Departure Time)
      - C7(AGENT), C8(VESSEL), I8(ROOM NO), C9(GUEST), I9(NATIONALITY)
      - 표는 14행부터(헤더는 13행), TOTAL은 I41 수식 그대로 유지

    ✅ 날짜별 출력 순서(표 시작 14행):
      - 체크인 날짜:  Accommodation → (Breakfast/Lunch/Dinner)
      - 중간 날짜:    Breakfast → Accommodation → Lunch → Dinner
      - 체크아웃 날짜: (Breakfast/Lunch/Dinner)만 (숙박 없음)
      - Early/Late/Extra night 규칙은 기존 계산 결과에 따라 숙박일(acc_dates)에 포함되면 숙박으로 찍힘
    """

    payer = next((x for x in stays_in_room if x.charge_room), stays_in_room[0])
    payer_rate = rates[(payer.agency_name, payer.room_type)]
    acc_desc = get_accommodation_base_desc(payer.agency_name)

    # ----- Header cells (value only) -----
    safe_set_addr(ws, "C7", agency_display_for_excel)  # Agent
    safe_set_addr(ws, "C8", vessel)                    # Vessel
    safe_set_addr(ws, "I8", int(room_no) if str(room_no).isdigit() else room_no)  # Room No(예: 512)

    # Guest names (Twin이면 2명 세로)
    crew_names = [s.crew_name for s in stays_in_room if (s.crew_name or "").strip()]
    safe_set_addr(ws, "C9", "\n".join(crew_names))
    safe_set_wrap_left(ws, "C9")

    # Nationality (세로)
    nat_lines = []
    for s in stays_in_room:
        n = normalize_multiline(getattr(s, "nationality", "") or "")
        if n:
            nat_lines.extend([x for x in n.split("\n") if x.strip()])
    safe_set_addr(ws, "I9", "\n".join(nat_lines))
    safe_set_wrap_left(ws, "I9")

    # --- 2명 이상일 때 Guest/Nationality 칸(9행) 높이 늘리기 ---
    try:
        name_lines = 1
        nat_lines = 1
        try:
            v_name = ws["C9"].value
            if isinstance(v_name, str) and v_name.strip():
                name_lines = max(1, v_name.count("\n") + 1)
        except Exception:
            pass
        try:
            v_nat = ws["I9"].value
            if isinstance(v_nat, str) and v_nat.strip():
                nat_lines = max(1, v_nat.count("\n") + 1)
        except Exception:
            pass
        lines = max(name_lines, nat_lines)
        if lines >= 2:
            # 기본 행높이(대략 15) * 줄수, 최소 30
            ws.row_dimensions[9].height = max(30, 15 * lines)
    except Exception:
        pass


    # Arrival/Departure
    safe_set_addr(ws, "C10", payer.checkin.date())
    safe_set_addr(ws, "C11", payer.checkin.strftime("%H:%M"))
    safe_set_addr(ws, "I10", payer.checkout.date())
    safe_set_addr(ws, "I11", payer.checkout.strftime("%H:%M"))

    # ----- Table range (value only) -----
    TABLE_START = 14
    TABLE_END = 40  # TOTAL이 41행에 있으므로 그 위까지만 사용
    COLS = (1, 3, 5, 9)  # A,C,E,I

    # 표 영역 value 초기화 (서식 유지)
    for r in range(TABLE_START, TABLE_END + 1):
        for c in COLS:
            cell = ws.cell(row=r, column=c)
            tl = _top_left_of_merged(ws, cell)
            if isinstance(tl, MergedCell):
                continue
            tl.value = None

    # 오른쪽(J열 이후) 값은 생성하지 않음(페이지 분할 방지)
    for r in range(1, 120):
        for c in range(10, 60):  # J ~
            cell = ws.cell(row=r, column=c)
            tl = _top_left_of_merged(ws, cell)
            if isinstance(tl, MergedCell):
                continue
            tl.value = None
    # 특히 J26 강제 비움
    try:
        ws["J26"].value = None
    except Exception:
        pass

    # ----- line writer -----
    cur_row = TABLE_START
    no = 1

    def add_line(dd, desc: str, amt: int):
        nonlocal cur_row, no
        if cur_row > TABLE_END:
            return
        # A(NO), C(DATE), E(DESC), I(AMOUNT)
        for (col, val) in [(1, no), (3, dd), (5, desc), (9, amt if amt else None)]:
            cell = ws.cell(row=cur_row, column=col)
            tl = _top_left_of_merged(ws, cell)
            if isinstance(tl, MergedCell):
                continue
            tl.value = val
        no += 1
        cur_row += 1

    # ----- compute dates -----
    base_nights, extra_out, has_late, extra_in, has_early = calc_room_parts(payer.checkin, payer.checkout)

    # 숙박일(날짜별 1박): 체크인 날짜부터 base_nights 만큼
    acc_dates = [payer.checkin.date() + timedelta(days=i) for i in range(max(1, base_nights))]
    # 14:01 이후 추가 1박은 체크아웃 날짜에 숙박 라인 추가
    if extra_out == 1:
        acc_dates.append(payer.checkout.date())

    # Early(이전 체크인: 1박 추가)은 체크인 날짜 숙박일에 포함되어 있을 것
    # Early fee(08:01~09:59)는 별도 라인(숙박 다음)
    early_fee = payer_rate.early_0801_0959 if (has_early and payer_rate.early_0801_0959 > 0) else 0
    late_fee = payer_rate.late_1201_1400 if (has_late and payer_rate.late_1201_1400 > 0) else 0

    # 식사(날짜별 건건) — 기본 대리점은 아침 포함
    meals_map = meal_amounts_by_date(stays_in_room, rates, include_breakfast=True)

    # 출력할 날짜: 숙박일 + 식사 선택일 + 체크아웃 날짜(식사 있을 수 있음)
    all_dates = sorted(set(acc_dates) | set(meals_map.keys()) | {payer.checkout.date()})

    first_day = payer.checkin.date()
    last_day = payer.checkout.date()

    for dd in all_dates:
        b_amt, l_amt, d_amt = meals_map.get(dd, (0, 0, 0))

        if dd == first_day:
            # 체크인 날짜: 숙박 먼저
            if dd in acc_dates:
                add_line(dd, acc_desc, payer_rate.nightly)
                if early_fee:
                    add_line(dd, "Accommodation Charge(Early check in)", early_fee)

            # 식사 (있을 때만)
            if b_amt:
                add_line(dd, "Café Food (Breakfast)", b_amt)
            if l_amt:
                add_line(dd, "Café Food (Lunch)", l_amt)
            if d_amt:
                add_line(dd, "Café Food (Dinner)", d_amt)

        elif dd == last_day:
            # 체크아웃 날짜: 식사만 (숙박 없음)
            if b_amt:
                add_line(dd, "Café Food (Breakfast)", b_amt)
            if l_amt:
                add_line(dd, "Café Food (Lunch)", l_amt)
            if d_amt:
                add_line(dd, "Café Food (Dinner)", d_amt)

        else:
            # 중간 날짜: 아침 → 숙박 → 점심 → 저녁
            if b_amt:
                add_line(dd, "Café Food (Breakfast)", b_amt)

            if dd in acc_dates:
                add_line(dd, acc_desc, payer_rate.nightly)

            if l_amt:
                add_line(dd, "Café Food (Lunch)", l_amt)
            if d_amt:
                add_line(dd, "Café Food (Dinner)", d_amt)

    # Late check out fee는 마지막에 추가
    if late_fee:
        add_line(payer.checkout.date(), "Accommodation Charge(Late check out)", late_fee)

    # ----- Print setup (1-page width) -----
    try:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_area = "A1:I49"
        try:
            ws.col_breaks.brk = []
            ws.row_breaks.brk = []
        except Exception:
            pass
    except Exception:
        pass


def _fill_room_sheet_hyupsung(
    ws,
    room_no: str,
    room_type: str,
    agency_display_for_excel: str,
    vessel: str,
    stays_in_room: List[CrewStay],
    rates: Dict[Tuple[str, str], RateRow]
):
    """
    협성 ROOM 템플릿 규칙:
      - Accommodation / Early / Late : 15행부터(15~21)
      - Meal only : 27행부터(27~40)
      - Accommodation/Meal SUMIF가 있으므로 설명 prefix 유지:
          * "Accommodation"
          * "Café Food"
    """
    # 상단(협성 ROOM 시트도 기본 위치가 동일하게 배치되어있음)
    safe_set_addr(ws, "C7", agency_display_for_excel)  # Agent name value
    safe_set_addr(ws, "C8", vessel)                    # Vessel name value
    safe_set_addr(ws, "I8", f"{room_no} - {roomtype_to_english(room_type)}")  # Room no

    # 이름/국적
    crew_names = []
    nationalities = []
    for s in stays_in_room:
        crew_names.append(s.crew_name.strip())
        nationalities.append((s.nationality or "").strip())

    safe_set_addr(ws, "C9", "\n".join([x for x in crew_names if x]))
    safe_set_wrap_left(ws, "C9")
    nat_lines = []
    for x in nationalities:
        x2 = normalize_multiline(x)
        if x2:
            nat_lines.extend(x2.split("\n"))
    safe_set_addr(ws, "I9", "\n".join([x for x in nat_lines if x]))
    safe_set_wrap_left(ws, "I9")

    # --- 2명 이상일 때 Guest/Nationality 칸(9행) 높이 늘리기 ---
    try:
        name_lines = 1
        nat_lines = 1
        try:
            v_name = ws["C9"].value
            if isinstance(v_name, str) and v_name.strip():
                name_lines = max(1, v_name.count("\n") + 1)
        except Exception:
            pass
        try:
            v_nat = ws["I9"].value
            if isinstance(v_nat, str) and v_nat.strip():
                nat_lines = max(1, v_nat.count("\n") + 1)
        except Exception:
            pass
        lines = max(name_lines, nat_lines)
        if lines >= 2:
            # 기본 행높이(대략 15) * 줄수, 최소 30
            ws.row_dimensions[9].height = max(30, 15 * lines)
    except Exception:
        pass


    # 날짜/시간
    min_checkin = min(s.checkin for s in stays_in_room)
    max_checkout = max(s.checkout for s in stays_in_room)
    safe_set_addr(ws, "C10", min_checkin.date())
    safe_set_addr(ws, "I10", max_checkout.date())
    safe_set_addr(ws, "C11", min_checkin.time())
    safe_set_addr(ws, "I11", max_checkout.time())

    # 1) Accommodation 영역 초기화(15~21: A,C,E,I 중심)
    _clear_range(ws, 15, 21, 1, 12)

    # 2) Meal 영역 초기화(27~40)
    _clear_range(ws, 27, 40, 1, 12)

    payer = next((x for x in stays_in_room if x.charge_room), stays_in_room[0])
    payer_rate = rates[(payer.agency_name, payer.room_type)]
    base_nights, extra_out, has_late, extra_in, has_early = calc_room_parts(payer.checkin, payer.checkout)
    total_nights = base_nights + extra_in + extra_out

    # ✅ 숙박은 날짜별로 1박씩 라인 생성
    acc_dates: List[date] = [payer.checkin.date() + timedelta(days=i) for i in range(base_nights)]
    if extra_in == 1:
        acc_dates.insert(0, payer.checkin.date())
    if extra_out == 1:
        acc_dates.append(payer.checkout.date())


    # ----- Accommodation lines in rows 15..21 -----
    acc_row = 15
    acc_no = 1

    def write_acc(desc: str, amt: int, d: date):
        nonlocal acc_row, acc_no
        if amt <= 0 or acc_row > 21:
            return
        safe_set_rc(ws, acc_row, 1, acc_no)  # A: NO
        safe_set_rc(ws, acc_row, 3, d)       # C: Date
        safe_set_rc(ws, acc_row, 5, desc)    # E: Description
        safe_set_rc(ws, acc_row, 9, amt)     # I: Amount
        acc_no += 1
        acc_row += 1

    # 첫 1박
    for dd in acc_dates:
        write_acc("Accommodation Charge(Breakfast included)", payer_rate.nightly, dd)

    # Early fee (08:01~09:59)
    if has_early and payer_rate.early_0801_0959 > 0:
        write_acc("Accommodation Charge(Early check in)", payer_rate.early_0801_0959, payer.checkin.date())

    # 나머지 숙박
    # (날짜별 숙박 라인 생성으로 remaining_nights 합산 라인 제거)

    # Late fee
    if has_late and payer_rate.late_1201_1400 > 0:
        write_acc("Accommodation Charge(Late check out)", payer_rate.late_1201_1400, payer.checkout.date())

    # ----- Meal lines in rows 27..40 -----
    meal_row = 27
    meal_no = 1

    def write_meal(desc: str, amt: int, d: date):
        nonlocal meal_row, meal_no
        if amt <= 0 or meal_row > 40:
            return
        safe_set_rc(ws, meal_row, 1, meal_no)  # A: NO
        safe_set_rc(ws, meal_row, 3, d)        # C: Date
        safe_set_rc(ws, meal_row, 5, desc)     # E: Description
        safe_set_rc(ws, meal_row, 9, amt)      # I: Amount
        meal_no += 1
        meal_row += 1
    # ✅ 협성: 조식은 숙박에 포함 → 식사(27행~)에는 점심/저녁만 날짜별로 기록
    meals_map = meal_amounts_by_date(stays_in_room, rates, include_breakfast=False)
    for dd in sorted(meals_map.keys()):
        _b, l_amt, d_amt = meals_map[dd]
        write_meal("Café Food (Lunch)", l_amt, dd)
        write_meal("Café Food (Dinner)", d_amt, dd)


# =========================
# Template selection & metadata
# =========================
def choose_template_for_stays(stays: List[CrewStay]) -> Path:
    """대리점별 템플릿 선택.
    ✅ 한글/영문 어떤 값이 들어와도 협성/동진만 예외 템플릿 사용.
    """
    agencies = sorted({(s.agency_name or "").strip() for s in stays})
    if len(agencies) != 1:
        return TEMPLATE_DEFAULT

    agency = agencies[0]

    # 1) 정확히 매칭
    if agency in TEMPLATE_EXCEPTIONS:
        return TEMPLATE_EXCEPTIONS[agency]

    # 2) 부분 매칭(협성*, *협성*, 동진*)
    if "협성" in agency:
        return TEMPLATE_EXCEPTIONS.get("협성", TEMPLATE_DEFAULT)
    if "동진" in agency:
        return TEMPLATE_EXCEPTIONS.get("동진", TEMPLATE_DEFAULT)

    return TEMPLATE_DEFAULT


def template_is_hyupsung(path: Path) -> bool:
    return path.name.lower() == "invoice_template_hyupsung.xlsx".lower()


def get_template_metadata(wb, tpl_path: Path) -> Tuple[str, str]:
    """
    returns (bill_sheet_name, base_room_sheet_name)
    협성 템플릿: BILL / ROOM
    기본 템플릿: TOTAL INVOICE (HOTEL) 또는 BILL / 402
    """
    if "BILL" in wb.sheetnames and "ROOM" in wb.sheetnames:
        return "BILL", "ROOM"

    bill = TOTAL_SHEET_NAME_DEFAULT if TOTAL_SHEET_NAME_DEFAULT in wb.sheetnames else ("BILL" if "BILL" in wb.sheetnames else "BILL")
    base = BASE_ROOM_SHEET_DEFAULT if BASE_ROOM_SHEET_DEFAULT in wb.sheetnames else ("ROOM" if "ROOM" in wb.sheetnames else BASE_ROOM_SHEET_DEFAULT)
    return bill, base


# =========================
# Save invoice
# =========================

def save_invoice_by_vessel(vessel: str, stays: List[CrewStay], rates: Dict[Tuple[str, str], RateRow]) -> Path:
    if not stays:
        raise ValueError("합산할 데이터가 없습니다.")

    agency_kr = (stays[0].agency_name or "").strip()
    if is_dongjin_agency(agency_kr):
        return _dongjin_save_invoice(vessel, stays, rates)

    tpl_path = choose_template_for_stays(stays)
    if not tpl_path.exists():
        raise FileNotFoundError(f"템플릿 파일이 없습니다: {tpl_path}")

    wb = load_workbook(tpl_path)
    rehydrate_images(wb)

    bill_sheet_name, base_room_sheet_name = get_template_metadata(wb, tpl_path)

    global BASE_ROOM_IMAGE_CACHE
    if base_room_sheet_name in wb.sheetnames:
        BASE_ROOM_IMAGE_CACHE = _build_image_cache(wb[base_room_sheet_name])
    else:
        BASE_ROOM_IMAGE_CACHE = []

    bill = wb[bill_sheet_name]

    agencies_kor = sorted({s.agency_name for s in stays})
    agency_disp_kor = agencies_kor[0] if len(agencies_kor) == 1 else "MIXED"
    agency_disp_eng = agency_display_eng_for_group(stays, rates, agency_disp_kor)

    min_checkin = min(s.checkin for s in stays)
    max_checkout = max(s.checkout for s in stays)

    unique_checkins = sorted({s.checkin for s in stays})
    if len(unique_checkins) == 1:
        checkin_dates_value = unique_checkins[0].date()
        checkin_times_value = unique_checkins[0].time()
    else:
        checkin_dates_value = "\n".join(format_date_dd_mmm_yy(dt.date()) for dt in unique_checkins)
        checkin_times_value = "\n".join(dt.strftime("%H:%M") for dt in unique_checkins)

    safe_set_addr(bill, "C7", agency_disp_eng)
    safe_set_addr(bill, "C8", vessel)
    safe_set_addr(bill, "C10", checkin_dates_value)
    safe_set_addr(bill, "C11", checkin_times_value)
    safe_set_wrap_left(bill, "C10")
    safe_set_wrap_left(bill, "C11")
    safe_set_addr(bill, "I10", max_checkout.date())
    safe_set_addr(bill, "I11", max_checkout.time())

    room_type_map = {}
    for s in stays:
        room_type_map[(s.room_no, s.room_type)] = s.room_type

    type_counter = {}
    for _, rt in room_type_map.items():
        eng = roomtype_to_english(rt)
        type_counter[eng] = type_counter.get(eng, 0) + 1

    parts = [f"{cnt} - {rtype}" for rtype, cnt in type_counter.items()]
    safe_set_addr(bill, "I8", " / ".join(parts))

    START_ROW = 14
    _clear_bill_table_cells(bill, START_ROW, START_ROW + 400)

    groups = defaultdict(list)
    for s in stays:
        groups[(s.room_no, s.checkin, s.checkout)].append(s)

    row_idx = 0
    grand_total = 0
    for key in sorted(groups.keys(), key=lambda k: (k[0], k[1], k[2])):
        room_no, checkin, checkout = key
        members = groups[key]

        guest_lines = [m.crew_name.strip() for m in members if m.crew_name.strip()]
        guest_name = "\n".join(guest_lines)

        payer = next((m for m in members if m.charge_room), members[0])
        payer_rate = rates[(payer.agency_name, payer.room_type)]

        room_total = late_fee = early_fee = 0
        if payer.charge_room:
            room_total, late_fee, early_fee = calc_room_charge(payer.checkin, payer.checkout, payer_rate)

        include_bf = True
        if any(("협성" in (m.agency_name or "")) or ("동진" in (m.agency_name or "")) for m in members):
            include_bf = False
        meals_map = meal_amounts_by_date(members, rates, include_breakfast=include_bf)
        meals_total = sum(v[0] + v[1] + v[2] for v in meals_map.values())

        line_total = room_total + late_fee + early_fee + meals_total
        grand_total += line_total

        row = START_ROW + row_idx
        row_idx += 1

        try:
            _copy_row_style(bill, START_ROW, row, 9)
        except Exception:
            pass

        safe_set_rc(bill, row, 1, row_idx)
        safe_set_rc(bill, row, 3, checkout.date())
        safe_set_rc(bill, row, 5, guest_name)
        bill.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        safe_set_rc(bill, row, 9, line_total)
        bill.cell(row=row, column=9).number_format = "#,##0"
        bill.row_dimensions[row].height = 30 if len(members) >= 2 else 18

    total_sum = 0
    for r in range(START_ROW, START_ROW + row_idx):
        val = bill.cell(row=r, column=9).value
        if isinstance(val, (int, float)):
            total_sum += val

    if bill_sheet_name == TOTAL_SHEET_NAME_DEFAULT:
        _set_total_invoice_default_total(bill)
    else:
        try:
            _set_total_amount_cells(bill, mode=('hyupsung' if template_is_hyupsung(tpl_path) else 'default'))
        except Exception:
            pass

    rooms: Dict[str, List[CrewStay]] = {}
    for s in stays:
        rooms.setdefault(s.room_no, []).append(s)

    for room_no, room_stays in rooms.items():
        sheet_name = _ensure_room_sheet(wb, room_no, base_room_sheet_name)
        ws_room = wb[sheet_name]

        rtypes = sorted({x.room_type for x in room_stays})
        room_type_disp = rtypes[0] if len(rtypes) == 1 else "Mixed"

        if template_is_hyupsung(tpl_path):
            _fill_room_sheet_hyupsung(
                ws_room, room_no, room_type_disp, agency_disp_eng, vessel, room_stays, rates
            )
        else:
            _fill_room_sheet_default(
                ws_room, room_no, room_type_disp, agency_disp_eng, vessel, room_stays, rates
            )

    if base_room_sheet_name == "ROOM" and "ROOM" in wb.sheetnames:
        try:
            wb.remove(wb["ROOM"])
        except Exception:
            pass

    rehydrate_images(wb)
    _apply_one_page_print_settings_to_workbook(wb)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    safe_vessel = "".join(ch for ch in vessel if ch.isalnum() or ch in ("-", "_")).strip() or "VESSEL"
    fname = f"INVOICE_{safe_vessel}_{min_checkin.strftime('%Y%m%d')}.xlsx"
    out_path = OUTPUT_DIR / fname

    attempt = 0
    while True:
        try:
            try:
                Path(OUTPUT_DIR).parent.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            wb.save(out_path)
            try:
                _ui_info("생성 완료", f"저장 완료!\n\n저장 파일: {Path(OUTPUT_DIR).resolve()}")
            except Exception:
                pass
            break
        except PermissionError:
            if not OUTPUT_DIR.exists():
                raise
            attempt += 1
            out_path = OUTPUT_DIR / f"{Path(fname).stem}_{attempt}.xlsx"
            if attempt >= 50:
                raise

    return out_path


# =========================
# Meal planner popup
# =========================
class MealPlanner(tk.Toplevel):
    def __init__(
        self,
        master: tk.Tk,
        title_suffix: str,
        dates: List[date],
        initial: Optional[Dict[str, Tuple[bool, bool, bool]]] = None
    ):
        super().__init__(master)
        self.title(f"식사 날짜별 선택 {title_suffix}".strip())

        self.geometry("760x520")
        self.minsize(900, 600)
        self.resizable(True, True)

        self._vars: Dict[str, Tuple[tk.BooleanVar, tk.BooleanVar, tk.BooleanVar]] = {}

        root = ttk.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        ttk.Label(root, text="날짜별로 아침/점심/저녁을 체크하세요").pack(anchor="w", pady=(0, 8))

        body = ttk.Frame(root)
        body.pack(fill="both", expand=True)

        left = ttk.Frame(body)
        left.pack(side="left", fill="both", expand=True)

        canvas = tk.Canvas(left)
        scrollbar = ttk.Scrollbar(left, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        right = ttk.Frame(body)
        right.pack(side="right", fill="y", padx=(12, 0))

        hdr = ttk.Frame(scroll_frame)
        hdr.pack(fill="x", pady=(0, 4))
        ttk.Label(hdr, text="Date", width=14).grid(row=0, column=0, sticky="w")
        ttk.Label(hdr, text="Breakfast").grid(row=0, column=1, padx=28, sticky="w")
        ttk.Label(hdr, text="Lunch").grid(row=0, column=2, padx=28, sticky="w")
        ttk.Label(hdr, text="Dinner").grid(row=0, column=3, padx=28, sticky="w")

        init = initial or {}
        for d in dates:
            d_str = d.strftime("%Y-%m-%d")
            b0, l0, d0 = init.get(d_str, (False, False, False))

            vb = tk.BooleanVar(value=b0)
            vl = tk.BooleanVar(value=l0)
            vd = tk.BooleanVar(value=d0)
            self._vars[d_str] = (vb, vl, vd)

            row = ttk.Frame(scroll_frame)
            row.pack(fill="x", pady=2)

            ttk.Label(row, text=d_str, width=14).grid(row=0, column=0, sticky="w")
            ttk.Checkbutton(row, variable=vb).grid(row=0, column=1, padx=40)
            ttk.Checkbutton(row, variable=vl).grid(row=0, column=2, padx=40)
            ttk.Checkbutton(row, variable=vd).grid(row=0, column=3, padx=40)

        # ✅ 버튼 세로 정렬 + 전체선택/해제
        ttk.Button(right, text="전체 선택", command=self._select_all).pack(fill="x", pady=(0, 6))
        ttk.Button(right, text="전체 해제", command=self._clear_all).pack(fill="x", pady=(0, 16))
        ttk.Button(right, text="저장", command=self._save).pack(fill="x", pady=(0, 6))
        ttk.Button(right, text="취소", command=self._cancel).pack(fill="x")

        self.result: Optional[Dict[str, Tuple[bool, bool, bool]]] = None
        self.grab_set()
        self.transient(master)

    def _select_all(self):
        for vb, vl, vd in self._vars.values():
            vb.set(True); vl.set(True); vd.set(True)

    def _clear_all(self):
        for vb, vl, vd in self._vars.values():
            vb.set(False); vl.set(False); vd.set(False)

    def _save(self):
        out: Dict[str, Tuple[bool, bool, bool]] = {}
        for d_str, (vb, vl, vd) in self._vars.items():
            out[d_str] = (bool(vb.get()), bool(vl.get()), bool(vd.get()))
        self.result = out
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()


# =========================
# GUI App
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()

        # ---- UI 스타일(깔끔하게) ----
        try:
            style = ttk.Style(self)
            try:
                style.theme_use("clam")
            except Exception:
                pass
            style.configure("TLabelframe", padding=10)
            style.configure("TLabelframe.Label", font=("Segoe UI", 11, "bold"))
            style.configure("TButton", padding=(8, 4))
            style.configure("TEntry", padding=(6, 4))
            # 위젯별 기본 폰트 크게
            try:
                style.configure("TLabel", font=("Segoe UI", 12))
                style.configure("TButton", font=("Segoe UI", 11, "bold"), padding=(12, 8))
                try:
                    style.configure("Action.TButton", font=("Segoe UI", 11, "bold"), padding=(14, 10))
                except Exception:
                    pass

                style.configure("TEntry", font=("Segoe UI", 12), padding=(8, 6))
                style.configure("TCombobox", font=("Segoe UI", 12), padding=(6, 4))
                style.configure("Treeview", font=("Segoe UI", 11), rowheight=24)
                style.configure("Treeview.Heading", font=("Segoe UI", 12, "bold"))
                style.configure("Big.TCheckbutton", font=("Segoe UI", 12, "bold"), padding=(6, 2))
            except Exception:
                pass

            # 화면 비율/가독성 개선(전체 스케일)
            try:
                # 기본 1.0 -> 1.15 (필요하면 1.10~1.25 조절)
                self.tk.call('tk', 'scaling', 1.20)
            except Exception:
                pass

            # 기본 폰트 크게(전체 UI 가독성)
            try:
                import tkinter.font as tkfont
                default_font = tkfont.nametofont("TkDefaultFont")
                text_font = tkfont.nametofont("TkTextFont")
                fixed_font = tkfont.nametofont("TkFixedFont")
                default_font.configure(size=12)
                text_font.configure(size=12)
                fixed_font.configure(size=11)
                self.option_add("*Font", default_font)
            except Exception:
                pass

            # 체크박스(식사) 더 크게 보이게(글자/패딩)
            try:
                style.configure("Big.TCheckbutton", font=("Segoe UI", 11), padding=(8, 6))
            except Exception:
                pass

        except Exception:
            pass

        self.title("호텔 인보이스 자동화")
        self.geometry("1520x860")

        self.rates: Dict[Tuple[str, str], RateRow] = load_rates(RATES_FILE)

        self.meal_maps: Dict[str, Dict[str, Tuple[bool, bool, bool]]] = {}
        self.current_meals_1: Dict[str, Tuple[bool, bool, bool]] = {}
        self.current_meals_2: Dict[str, Tuple[bool, bool, bool]] = {}

        self._build_ui()
        self._refresh_agency_list()



    def _normalize_dt_var(self, var: tk.StringVar) -> None:
        """입력창에서 날짜/시간을 타이핑해도 yyyy-mm-dd HH:MM(10분단위)로 정규화."""
        try:
            d, hh, mm = self._dt_parse(var.get())
            var.set(f"{d} {hh}:{mm}")
        except Exception:
            pass

    def _dt_parse(self, s: str):
        s = (s or "").strip()
        d = datetime.now().strftime("%Y-%m-%d")
        hh = "10"
        mm = "00"
        try:
            parts = s.split()
            if len(parts) >= 1 and parts[0]:
                d = parts[0]
            if len(parts) >= 2 and ":" in parts[1]:
                hh, mm = parts[1].split(":")[:2]
                hh = hh.zfill(2)
                mm = mm.zfill(2)
        except Exception:
            pass
        allowed = {"00","10","20","30","40","50"}
        if mm not in allowed:
            mm = "00"
        try:
            ih = int(hh)
            if ih < 0 or ih > 23:
                hh = "10"
        except Exception:
            hh = "10"
        return d, hh, mm

    def _make_inline_dt_panel(self, parent, target_var: tk.StringVar, title: str):
        # 내장 달력(개선판): 오늘/주말/선택 강조 + 월/년 빠른 이동 + 키보드 조작 + Esc 닫기 + Today 버튼
        panel = ttk.Labelframe(parent, text=title, padding=(8, 6))
        panel.columnconfigure(0, weight=1)
        try:
            panel.rowconfigure(1, weight=1)
        except Exception:
            pass

        d0, h0, m0 = self._dt_parse(target_var.get())
        date_var = tk.StringVar(value=d0)
        hour_var = tk.StringVar(value=h0)
        min_var = tk.StringVar(value=m0)

        def apply_now():
            d = date_var.get().strip()
            hh = hour_var.get().strip().zfill(2)
            mm = min_var.get().strip().zfill(2)
            allowed = {"00","10","20","30","40","50"}
            if mm not in allowed:
                mm = "00"
                min_var.set(mm)
            try:
                ih = int(hh)
                if ih < 0 or ih > 23:
                    hh = "10"
                    hour_var.set(hh)
            except Exception:
                hh = "10"
                hour_var.set(hh)
            target_var.set(f"{d} {hh}:{mm}")

        # ---- state ----
        today = date.today()
        try:
            y, mo, dd = [int(x) for x in date_var.get().split("-")]
            selected = date(y, mo, dd)
        except Exception:
            selected = today

        state = {
            "year": selected.year,
            "month": selected.month,
            "selected": selected,
        }

        # ---- header: month/year + navigation ----
        header = ttk.Frame(panel)
        header.grid(row=0, column=0, columnspan=7, sticky="we")
        header.columnconfigure(3, weight=1)

        year_var = tk.IntVar(value=state["year"])
        month_var = tk.IntVar(value=state["month"])

        def clamp_year(yv: int) -> int:
            if yv < 2000:
                return 2000
            if yv > 2100:
                return 2100
            return yv

        def sync_state_from_vars():
            state["year"] = clamp_year(int(year_var.get()))
            mv = int(month_var.get())
            if mv < 1:
                mv = 1
            if mv > 12:
                mv = 12
            month_var.set(mv)
            state["month"] = mv

        def prev_month():
            sync_state_from_vars()
            mv = state["month"] - 1
            yv = state["year"]
            if mv <= 0:
                mv = 12
                yv -= 1
            year_var.set(clamp_year(yv))
            month_var.set(mv)
            try:
                cb_year.set(str(int(year_var.get())))
                cb_month.set(str(int(month_var.get())))
            except Exception:
                pass
            sync_state_from_vars()
            redraw()

        def next_month():
            sync_state_from_vars()
            mv = state["month"] + 1
            yv = state["year"]
            if mv >= 13:
                mv = 1
                yv += 1
            year_var.set(clamp_year(yv))
            month_var.set(mv)
            try:
                cb_year.set(str(int(year_var.get())))
                cb_month.set(str(int(month_var.get())))
            except Exception:
                pass
            sync_state_from_vars()
            redraw()

        ttk.Button(header, text="◀", width=3, command=prev_month).grid(row=0, column=0, sticky="w")

        years = [str(y) for y in range(2000, 2101)]
        months = [str(m) for m in range(1, 13)]

        cb_year = ttk.Combobox(header, values=years, width=6)
        cb_year.grid(row=0, column=1, padx=(6, 4), sticky="w")
        cb_year.set(str(state["year"]))

        cb_month = ttk.Combobox(header, values=months, width=4)
        cb_month.grid(row=0, column=2, padx=(0, 6), sticky="w")
        cb_month.set(str(state["month"]))

        ttk.Label(header, text="").grid(row=0, column=3, sticky="we")

        def on_year_month_change(_evt=None):
            try:
                year_var.set(int(cb_year.get()))
            except Exception:
                year_var.set(state["year"])
                cb_year.set(str(state["year"]))
            try:
                month_var.set(int(cb_month.get()))
            except Exception:
                month_var.set(state["month"])
                cb_month.set(str(state["month"]))
            sync_state_from_vars()
            redraw()

        cb_year.bind("<<ComboboxSelected>>", on_year_month_change)
        cb_month.bind("<<ComboboxSelected>>", on_year_month_change)
        cb_year.bind("<Return>", on_year_month_change)
        cb_month.bind("<Return>", on_year_month_change)

        ttk.Button(header, text="▶", width=3, command=next_month).grid(row=0, column=4, sticky="e")

        def go_today():
            state["selected"] = today
            year_var.set(today.year)
            month_var.set(today.month)
            cb_year.set(str(today.year))
            cb_month.set(str(today.month))
            sync_state_from_vars()
            set_selected(today, keep_month=True)

        ttk.Button(header, text="Today", width=6, command=go_today).grid(row=0, column=5, padx=(6, 0), sticky="e")

        # ---- calendar body ----
        cal_body = ttk.Frame(panel)
        cal_body.grid(row=1, column=0, columnspan=7, sticky="nsew", pady=(6, 0))
        for c in range(7):
            try:
                cal_body.columnconfigure(c, weight=1, uniform="cal")
            except Exception:
                pass
        for r in range(7):
            try:
                cal_body.rowconfigure(r, weight=1, uniform="calr")
            except Exception:
                pass


        # store day buttons by date
        day_btns = {}

        def set_selected(d: date, keep_month: bool = False):
            state["selected"] = d
            date_var.set(f"{d.year:04d}-{d.month:02d}-{d.day:02d}")
            apply_now()
            if not keep_month:
                year_var.set(d.year)
                month_var.set(d.month)
                cb_year.set(str(d.year))
                cb_month.set(str(d.month))
                sync_state_from_vars()
            redraw()

        def style_btn(btn, d: date, weekday_col: int):
            # base
            is_weekend = weekday_col in (0, 6)
            is_today = (d == today)
            is_sel = (d == state["selected"])

            # tk.Button styling
            try:
                btn.configure(relief=("sunken" if is_sel else "raised"))
            except Exception:
                pass
            # fg for weekend
            try:
                if is_weekend:
                    btn.configure(fg="#C00000" if weekday_col == 0 else "#0050B0")
                else:
                    btn.configure(fg="black")
            except Exception:
                pass
            # font emphasis
            try:
                if is_today and is_sel:
                    btn.configure(font=("Segoe UI", 10, "bold"))
                elif is_today:
                    btn.configure(font=("Segoe UI", 10, "bold"))
                elif is_sel:
                    btn.configure(font=("Segoe UI", 10, "bold"))
                else:
                    btn.configure(font=("Segoe UI", 10, "normal"))
            except Exception:
                pass

        def redraw():
            for w in cal_body.winfo_children():
                w.destroy()
            day_btns.clear()

            yv = state["year"]
            mv = state["month"]

            # weekday header
            wdays = ["Sun","Mon","Tue","Wed","Thu","Fri","Sat"]
            for c, name in enumerate(wdays):
                ttk.Label(cal_body, text=name, font=("Segoe UI", 10, "bold")).grid(row=0, column=c, padx=2, pady=(0, 2), sticky="nsew")

            cal = calendar.Calendar(firstweekday=6)
            weeks = cal.monthdayscalendar(yv, mv)

            for r, week in enumerate(weeks, start=1):
                for c, day in enumerate(week):
                    if day == 0:
                        ttk.Label(cal_body, text="").grid(row=r, column=c, padx=1, pady=1, sticky="nsew")
                        continue
                    d = date(yv, mv, day)
                    # tk.Button gives better control for colors/relief
                    btn = tk.Button(
                        cal_body,
                        text=str(day),
                        width=3,
                        padx=1,
                        pady=1,
                        height=1,
                        
                        command=lambda dd=d: set_selected(dd)
                    )
                    style_btn(btn, d, c)
                    btn.grid(row=r, column=c, padx=1, pady=1, sticky='nsew')
                    day_btns[d] = btn

            # focus for keyboard
            try:
                panel.focus_set()
            except Exception:
                pass

        redraw()

        # ---- keyboard navigation ----
        def move_days(delta: int):
            d = state["selected"] + timedelta(days=delta)
            set_selected(d)

        def on_key(event):
            k = event.keysym
            if k == "Left":
                move_days(-1)
            elif k == "Right":
                move_days(1)
            elif k == "Up":
                move_days(-7)
            elif k == "Down":
                move_days(7)
            elif k in ("Return", "KP_Enter"):
                apply_now()
            elif k == "Escape":
                try:
                    close_btn.invoke()
                except Exception:
                    pass
            return "break"

        panel.bind("<Left>", on_key)
        panel.bind("<Right>", on_key)
        panel.bind("<Up>", on_key)
        panel.bind("<Down>", on_key)
        panel.bind("<Return>", on_key)
        panel.bind("<KP_Enter>", on_key)
        panel.bind("<Escape>", on_key)

        ttk.Separator(panel, orient="horizontal").grid(row=2, column=0, columnspan=7, sticky="we", pady=(4, 4))

        # ---- time row ----
        time_row = ttk.Frame(panel)
        time_row.grid(row=3, column=0, columnspan=7, sticky="we")
        ttk.Label(time_row, text="Time").pack(side="left")

        hours = [str(i).zfill(2) for i in range(24)]
        mins = ["00","10","20","30","40","50"]

        cb_h = ttk.Combobox(time_row, textvariable=hour_var, values=hours, width=4)
        cb_h.pack(side="left", padx=(8, 0))
        ttk.Label(time_row, text=":").pack(side="left", padx=(4, 4))
        cb_m = ttk.Combobox(time_row, textvariable=min_var, values=mins, width=4)
        cb_m.pack(side="left")

        cb_h.bind("<<ComboboxSelected>>", lambda e: apply_now())
        cb_m.bind("<<ComboboxSelected>>", lambda e: apply_now())
        cb_h.bind("<KeyRelease>", lambda e: apply_now())
        cb_m.bind("<KeyRelease>", lambda e: apply_now())

        # ---- buttons ----
        btn_row = ttk.Frame(panel)
        btn_row.grid(row=4, column=0, columnspan=7, sticky="e", pady=(6, 0))
        close_btn = ttk.Button(btn_row, text="닫기")
        close_btn.pack(side="right")

        apply_now()
        return panel, close_btn



    def _refresh_window_size(self) -> None:
        """패널 열고/닫을 때 창이 내용에 맞게 다시 줄어들도록(강제)."""
        try:
            self.update_idletasks()
            # 최소 크기 제한이 있으면 줄어들지 않으므로 리셋
            try:
                self.minsize(900, 600)
            except Exception:
                pass
            # 요청 크기로 강제 설정(geometry("")만으로는 줄어들지 않는 케이스 대응)
            w = self.winfo_reqwidth()
            h = self.winfo_reqheight()
            # 화면 밖으로 나가지 않게 최소/최대 보호
            if w < 300:
                w = 300
            if h < 200:
                h = 200
            self.geometry(f"{w}x{h}")
            self.update_idletasks()
        except Exception:
            pass
    def _hide_holder(self, holder: ttk.Frame) -> None:
        try:
            holder.grid_remove()
        except Exception:
            pass
        self._refresh_window_size()

    def _toggle_holder(self, holder: ttk.Frame, other: ttk.Frame | None = None) -> None:
        try:
            if other is not None and other.winfo_ismapped():
                other.grid_remove()
        except Exception:
            pass
        try:
            if holder.winfo_ismapped():
                holder.grid_remove()
            else:
                holder.grid()
        except Exception:
            pass
        self._refresh_window_size()

    def _toggle_panel(self, panel: ttk.Frame, other=None) -> None:
        try:
            if other is not None and other.winfo_ismapped():
                other.grid_remove()
        except Exception:
            pass
        if panel.winfo_ismapped():
            panel.grid_remove()
        else:
            panel.grid()
        self._refresh_window_size()


    def _build_ui(self):
        top = ttk.Frame(self, padding=12)
        top.pack(fill="x")

        left = ttk.Frame(top)
        left.pack(side="left", fill="x", expand=True)

        actions = ttk.LabelFrame(top, text="작업", padding=10)
        actions.pack(side="right", fill="y", padx=(12, 0))
        try:
            actions.configure(width=180)
            actions.pack_propagate(False)
        except Exception:
            pass

        basic = ttk.LabelFrame(left, text="기본 정보", padding=10)
        basic.pack(fill="x")
        for i in range(6):
            basic.columnconfigure(i, weight=1)

        ttk.Label(basic, text="선박명").grid(row=0, column=0, sticky="w")
        self.vessel_var = tk.StringVar()
        ttk.Entry(basic, textvariable=self.vessel_var, width=22).grid(row=0, column=1, sticky="we", padx=(6, 14), pady=4)

        ttk.Label(basic, text="룸 번호").grid(row=0, column=2, sticky="w")
        self.roomno_var = tk.StringVar()
        ttk.Entry(basic, textvariable=self.roomno_var, width=12).grid(row=0, column=3, sticky="we", padx=(6, 14), pady=4)

        ttk.Label(basic, text="대리점(한글)").grid(row=0, column=4, sticky="w")
        self.agency_var = tk.StringVar()
        self.agency_cb = ttk.Combobox(basic, textvariable=self.agency_var, state="readonly", width=18)
        self.agency_cb.grid(row=0, column=5, sticky="we", padx=(6, 0), pady=4)
        self.agency_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_room_types())

        ttk.Label(basic, text="룸타입").grid(row=1, column=0, sticky="w")
        self.roomtype_var = tk.StringVar()
        self.roomtype_cb = ttk.Combobox(basic, textvariable=self.roomtype_var, state="readonly", width=18)
        self.roomtype_cb.grid(row=1, column=1, sticky="we", padx=(6, 14), pady=4)
        self.roomtype_cb.bind("<<ComboboxSelected>>", self._roomtype_changed)

        today_str = datetime.now().strftime("%Y-%m-%d")

        ttk.Label(basic, text="입실(24h)").grid(row=1, column=2, sticky="w")
        self.checkin_var = tk.StringVar(value=f"{today_str} 10:00")
        ttk.Entry(basic, textvariable=self.checkin_var, width=18).grid(row=1, column=3, sticky="we", padx=(6, 14), pady=4)
        try:
            self.checkin_entry = basic.grid_slaves(row=1, column=3)[0]
            self.checkin_entry.bind('<FocusOut>', lambda e, v=self.checkin_var: self._normalize_dt_var(v))
        except Exception:
            pass
        btn_in = ttk.Button(basic, text="📅", width=3)
        btn_in.grid(row=1, column=3, sticky="e", padx=(0, 18), pady=4)

        self._checkin_panel_holder = ttk.Frame(basic)
        self._checkin_panel_holder.grid(row=2, column=3, sticky="w", padx=(6, 14))
        self.checkin_panel, close_in = self._make_inline_dt_panel(self._checkin_panel_holder, self.checkin_var, "입실 선택")
        self.checkin_panel.grid(row=0, column=0, sticky="w")
        self._checkin_panel_holder.grid_remove()
        btn_in.configure(command=lambda: self._toggle_holder(self._checkin_panel_holder, getattr(self, "_checkout_panel_holder", None)))
        close_in.configure(command=lambda: self._hide_holder(self._checkin_panel_holder))

        ttk.Label(basic, text="퇴실(24h)").grid(row=1, column=4, sticky="w")
        self.checkout_var = tk.StringVar(value=f"{today_str} 14:00")
        ttk.Entry(basic, textvariable=self.checkout_var, width=18).grid(row=1, column=5, sticky="we", padx=(6, 0), pady=4)
        try:
            self.checkout_entry = basic.grid_slaves(row=1, column=5)[0]
            self.checkout_entry.bind('<FocusOut>', lambda e, v=self.checkout_var: self._normalize_dt_var(v))
        except Exception:
            pass
        btn_out = ttk.Button(basic, text="📅", width=3)
        btn_out.grid(row=1, column=5, sticky="e", padx=(0, 6), pady=4)

        self._checkout_panel_holder = ttk.Frame(basic)
        self._checkout_panel_holder.grid(row=2, column=5, sticky="w", padx=(6, 0))
        self.checkout_panel, close_out = self._make_inline_dt_panel(self._checkout_panel_holder, self.checkout_var, "퇴실 선택")
        self.checkout_panel.grid(row=0, column=0, sticky="w")
        self._checkout_panel_holder.grid_remove()
        btn_out.configure(command=lambda: self._toggle_holder(self._checkout_panel_holder, getattr(self, "_checkin_panel_holder", None)))
        close_out.configure(command=lambda: self._hide_holder(self._checkout_panel_holder))

        guests = ttk.LabelFrame(left, text="선원", padding=10)
        guests.pack(fill="x", pady=(10, 0))
        for i in range(6):
            guests.columnconfigure(i, weight=1)

        ttk.Label(guests, text="선원명1").grid(row=0, column=0, sticky="w")
        self.crew1_var = tk.StringVar()
        ttk.Entry(guests, textvariable=self.crew1_var).grid(row=0, column=1, sticky="we", padx=(6, 14), pady=4)

        ttk.Label(guests, text="국적1").grid(row=0, column=2, sticky="w")
        self.nat1_var = tk.StringVar()
        ttk.Entry(guests, textvariable=self.nat1_var, width=10).grid(row=0, column=3, sticky="we", padx=(6, 14), pady=4)

        self.crew2_label = ttk.Label(guests, text="선원명2 (Twin)")
        self.crew2_label.grid(row=1, column=0, sticky="w")
        self.crew2_var = tk.StringVar()
        self.crew2_entry = ttk.Entry(guests, textvariable=self.crew2_var)
        self.crew2_entry.grid(row=1, column=1, sticky="we", padx=(6, 14), pady=4)

        self.nat2_label = ttk.Label(guests, text="국적2 (Twin)")
        self.nat2_label.grid(row=1, column=2, sticky="w")
        self.nat2_var = tk.StringVar()
        self.nat2_entry = ttk.Entry(guests, textvariable=self.nat2_var, width=10)
        self.nat2_entry.grid(row=1, column=3, sticky="we", padx=(6, 14), pady=4)

        # 기본 숨김
        self.crew2_label.grid_remove()
        self.crew2_entry.grid_remove()
        self.nat2_label.grid_remove()
        self.nat2_entry.grid_remove()

        meals = ttk.LabelFrame(left, text="식사", padding=10)
        meals.pack(fill="x", pady=(10, 0))
        meals.columnconfigure(0, weight=1)
        meals.columnconfigure(1, weight=1)

        m1 = ttk.Frame(meals)
        m1.grid(row=0, column=0, sticky="we", padx=(0, 10))
        self.meal1_summary_var = tk.StringVar(value="선원1 식사: (선택 없음)")
        ttk.Label(m1, textvariable=self.meal1_summary_var).pack(anchor="w")
        m1_btns = ttk.Frame(m1)
        m1_btns.pack(anchor="w", pady=(6, 0))
        ttk.Button(m1_btns, text="선원1 식사 선택", command=self._open_meal_planner_1).pack(side="left")
        ttk.Button(m1_btns, text="초기화", command=self._reset_meals_1).pack(side="left", padx=(6, 0))

        m2 = ttk.Frame(meals)
        m2.grid(row=0, column=1, sticky="we")
        self.meal2_summary_var = tk.StringVar(value="선원2 식사: (선택 없음)")
        self.meal2_label = ttk.Label(m2, textvariable=self.meal2_summary_var)
        self.meal2_label.pack(anchor="w")
        m2_btns = ttk.Frame(m2)
        m2_btns.pack(anchor="w", pady=(6, 0))
        ttk.Button(m2_btns, text="선원2 식사 선택", command=self._open_meal_planner_2).pack(side="left")
        ttk.Button(m2_btns, text="초기화", command=self._reset_meals_2).pack(side="left", padx=(6, 0))

        # Twin 아닐 때 숨김
        self.meal2_label.pack_forget()
        m2_btns.pack_forget()
        self._meal2_btns = m2_btns

        ttk.Button(actions, text="요율 새로고침", command=self._reload_rates).pack(fill="x", pady=3)
        ttk.Button(actions, text="초기화", command=self._reset_everything).pack(fill="x", pady=(3, 10))
        ttk.Button(actions, text="목록에 추가", command=self._add_row).pack(fill="x", pady=3)
        ttk.Button(actions, text="선택 행 삭제", command=self._delete_selected).pack(fill="x", pady=3)
        ttk.Separator(actions).pack(fill="x", pady=10)
        ttk.Button(actions, text="선택(선박별) 생성", command=self._generate_selected_group).pack(fill="x", pady=3)
        ttk.Button(actions, text="전체(선박별) 생성", command=self._generate_all_group).pack(fill="x", pady=3)

        table_frame = ttk.LabelFrame(self, text="목록", padding=10)
        table_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        columns = ("vessel", "roomno", "crew", "nat", "agency", "roomtype", "charge_room", "checkin", "checkout", "meals")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=18)
        self.tree.pack(side="left", fill="both", expand=True)

        headings = {
            "vessel": "선박",
            "roomno": "룸번호",
            "crew": "선원",
            "nat": "국적",
            "agency": "대리점(한글)",
            "roomtype": "룸타입",
            "charge_room": "룸요금",
            "checkin": "입실(24h)",
            "checkout": "퇴실(24h)",
            "meals": "식사",
        }
        for c in columns:
            self.tree.heading(c, text=headings[c])
            if c == "meals":
                self.tree.column(c, width=120, anchor="w")
            elif c in ("checkin", "checkout"):
                self.tree.column(c, width=160, anchor="w")
            elif c == "roomno":
                self.tree.column(c, width=80, anchor="w")
            elif c == "charge_room":
                self.tree.column(c, width=70, anchor="center")
            elif c == "nat":
                self.tree.column(c, width=80, anchor="w")
            else:
                self.tree.column(c, width=140, anchor="w")

        scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scroll.set)
        scroll.pack(side="right", fill="y")

        self.status_var = tk.StringVar(value="준비 완료!")
        ttk.Label(self, textvariable=self.status_var, padding=(12, 0, 12, 10)).pack(fill="x")

    def _roomtype_changed(self, event=None):
        twin = is_twin(self.roomtype_var.get())
        if twin:
            self.crew2_label.grid()
            self.crew2_entry.grid()
            self.nat2_label.grid()
            self.nat2_entry.grid()
            self.meal2_label.pack(anchor="w")
            self._meal2_btns.pack(anchor="w", pady=(6, 0))
        else:
            self.crew2_var.set("")
            self.nat2_var.set("")
            self._reset_meals_2()
            self.crew2_label.grid_remove()
            self.crew2_entry.grid_remove()
            self.nat2_label.grid_remove()
            self.nat2_entry.grid_remove()
            self.meal2_label.pack_forget()
            self._meal2_btns.pack_forget()


    def _reset_everything(self):
        """전체 입력/목록/식사 선택까지 초기화."""
        try:
            # 기본 입력값
            for v in [self.vessel_var, self.roomno_var, self.agency_var, self.roomtype_var,
                      self.crew1_var, self.crew2_var, self.nat1_var, self.nat2_var,
                      self.checkin_var, self.checkout_var]:
                try:
                    v.set("")
                except Exception:
                    pass

            # 기본값(선택형) 세팅
            try:
                if self.roomtype_cb is not None:
                    self.roomtype_cb.current(0)
            except Exception:
                pass
            try:
                if self.agency_cb is not None and len(self.agency_cb["values"]) > 0:
                    self.agency_cb.current(0)
            except Exception:
                pass

            # 달력 패널 닫기 + 창 크기 복원
            for holder_name in ["_checkin_panel_holder", "_checkout_panel_holder"]:
                holder = getattr(self, holder_name, None)
                if holder is not None:
                    try:
                        holder.grid_remove()
                    except Exception:
                        pass
            try:
                self._refresh_window_size()
            except Exception:
                pass

            # 식사 선택(현재 입력분) 초기화
            try:
                self.current_meals_1 = {}
                self.current_meals_2 = {}
                self.meal1_summary_var.set("")
                self.meal2_summary_var.set("")
            except Exception:
                pass

            # 트리 목록 초기화
            try:
                for iid in self.tree.get_children(""):
                    self.tree.delete(iid)
            except Exception:
                pass
            try:
                self.meal_maps = {}
            except Exception:
                pass

            self.status_var.set("전체 초기화 완료")
        except Exception as e:
            try:
                self.status_var.set(f"초기화 실패: {e}")
            except Exception:
                pass

    def _reload_rates(self):
        try:
            self.rates = load_rates(RATES_FILE)
            self._refresh_agency_list()
            self.status_var.set("요율을 다시 불러왔습니다!")
        except Exception as e:
            messagebox.showerror("새로고침 실패", str(e))

    def _refresh_agency_list(self):
        agencies = available_agencies(self.rates)
        self.agency_cb["values"] = agencies
        if agencies:
            if self.agency_var.get() not in agencies:
                self.agency_var.set(agencies[0])
            self._refresh_room_types()

    def _refresh_room_types(self):
        agency = self.agency_var.get().strip()
        rooms = available_room_types_for_agency(self.rates, agency)
        self.roomtype_cb["values"] = rooms
        if rooms and self.roomtype_var.get() not in rooms:
            self.roomtype_var.set(rooms[0])
        self._roomtype_changed()

    def _reset_meals_1(self):
        self.current_meals_1 = {}
        self.meal1_summary_var.set("선원1 식사: (선택 없음)")

    def _reset_meals_2(self):
        self.current_meals_2 = {}
        self.meal2_summary_var.set("선원2 식사: (선택 없음)")

    def _calc_meal_summary(self, meals: Dict[str, Tuple[bool, bool, bool]]) -> str:
        b = l = d = 0
        for _, (bb, ll, dd) in meals.items():
            b += 1 if bb else 0
            l += 1 if ll else 0
            d += 1 if dd else 0
        if b == 0 and l == 0 and d == 0:
            return "(선택 없음)"
        return f"조식 {b} / 중식 {l} / 석식 {d}"

    def _open_meal_planner_common(self, who: str, current: Dict[str, Tuple[bool, bool, bool]]) -> Optional[Dict[str, Tuple[bool, bool, bool]]]:
        try:
            checkin = parse_dt_24h(self.checkin_var.get())
            checkout = parse_dt_24h(self.checkout_var.get())
            if checkout < checkin:
                raise ValueError("퇴실일시는 입실일시보다 이후여야 합니다.")
        except Exception as e:
            messagebox.showerror("식사 선택 불가", f"입/퇴실 입력을 먼저 확인해 주세요.\n\n{e}")
            return None

        dates = daterange_inclusive(checkin.date(), checkout.date())
        planner = MealPlanner(self, f"({who})", dates, initial=current)
        self.wait_window(planner)
        return planner.result

    def _open_meal_planner_1(self):
        res = self._open_meal_planner_common("선원1", self.current_meals_1)
        if res is None:
            return
        self.current_meals_1 = res
        self.meal1_summary_var.set(f"선원1 식사: {self._calc_meal_summary(self.current_meals_1)}")

    def _open_meal_planner_2(self):
        res = self._open_meal_planner_common("선원2", self.current_meals_2)
        if res is None:
            return
        self.current_meals_2 = res
        self.meal2_summary_var.set(f"선원2 식사: {self._calc_meal_summary(self.current_meals_2)}")

    def _meals_text_for_tree(self, meals_by_date: Dict[str, Tuple[bool, bool, bool]]) -> str:
        b = l = d = 0
        for _, (bb, ll, dd) in meals_by_date.items():
            b += 1 if bb else 0
            l += 1 if ll else 0
            d += 1 if dd else 0
        return "없음" if (b == 0 and l == 0 and d == 0) else f"조{b}/중{l}/석{d}"

    def _validate_base_inputs(self) -> Tuple[str, str, str, str, str, str, datetime, datetime]:
        vessel = self.vessel_var.get().strip()
        room_no = self.roomno_var.get().strip()
        agency = self.agency_var.get().strip()
        room_type = self.roomtype_var.get().strip()

        crew1 = self.crew1_var.get().strip()
        nat1 = self.nat1_var.get().strip()

        if not vessel:
            raise ValueError("선박명을 입력해 주세요.")
        if not room_no:
            raise ValueError("룸 번호를 입력해 주세요.")
        if not agency:
            raise ValueError("대리점을 선택해 주세요.")
        if not room_type:
            raise ValueError("룸타입을 선택해 주세요.")
        if (agency, room_type) not in self.rates:
            raise ValueError(f"요율에 '{agency} / {room_type}' 조합이 없습니다.")
        if not crew1:
            raise ValueError("선원명1을 입력해 주세요.")

        checkin = parse_dt_24h(self.checkin_var.get())
        checkout = parse_dt_24h(self.checkout_var.get())
        if checkout <= checkin:
            raise ValueError("퇴실일시는 입실일시보다 이후여야 합니다.")

        return vessel, room_no, agency, room_type, crew1, nat1, checkin, checkout

    def _add_one_stay_to_tree(self, stay: CrewStay):
        iid = self.tree.insert("", "end", values=(
            stay.vessel,
            stay.room_no,
            stay.crew_name,
            stay.nationality,
            stay.agency_name,
            stay.room_type,
            "Y" if stay.charge_room else "N",
            stay.checkin.strftime("%Y-%m-%d %H:%M"),
            stay.checkout.strftime("%Y-%m-%d %H:%M"),
            self._meals_text_for_tree(stay.meals_by_date),
        ))
        self.meal_maps[iid] = stay.meals_by_date

    def _add_row(self):
        try:
            vessel, room_no, agency, room_type, crew1, nat1, checkin, checkout = self._validate_base_inputs()

            if is_twin(room_type):
                crew2 = self.crew2_var.get().strip()
                nat2 = self.nat2_var.get().strip()
                if not crew2:
                    raise ValueError("Twin(트윈) 선택 시 선원명2를 입력해 주세요.")

                stay1 = CrewStay(
                    vessel=vessel, room_no=room_no, crew_name=crew1, nationality=nat1,
                    agency_name=agency, room_type=room_type,
                    checkin=checkin, checkout=checkout,
                    meals_by_date=dict(self.current_meals_1),
                    charge_room=True
                )
                stay2 = CrewStay(
                    vessel=vessel, room_no=room_no, crew_name=crew2, nationality=nat2,
                    agency_name=agency, room_type=room_type,
                    checkin=checkin, checkout=checkout,
                    meals_by_date=dict(self.current_meals_2),
                    charge_room=False
                )
                self._add_one_stay_to_tree(stay1)
                self._add_one_stay_to_tree(stay2)
            else:
                stay = CrewStay(
                    vessel=vessel, room_no=room_no, crew_name=crew1, nationality=nat1,
                    agency_name=agency, room_type=room_type,
                    checkin=checkin, checkout=checkout,
                    meals_by_date=dict(self.current_meals_1),
                    charge_room=True
                )
                self._add_one_stay_to_tree(stay)

            self.status_var.set("목록에 추가했습니다!")
            self._reset_meals_1()
            self._reset_meals_2()
        except Exception as e:
            messagebox.showerror("추가 실패", str(e))

    def _delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("안내", "삭제할 행을 선택해 주세요.")
            return
        for iid in selected:
            self.tree.delete(iid)
            self.meal_maps.pop(iid, None)
        self.status_var.set("선택 행을 삭제했습니다!")

    def _row_to_stay(self, iid: str) -> CrewStay:
        vessel, roomno, crew, nat, agency, roomtype, charge_room, checkin_s, checkout_s, _ = self.tree.item(iid, "values")
        meals_map = self.meal_maps.get(iid, {})
        return CrewStay(
            vessel=str(vessel),
            room_no=str(roomno),
            crew_name=str(crew),
            nationality=str(nat),
            agency_name=str(agency),
            room_type=str(roomtype),
            checkin=parse_dt_24h(str(checkin_s)),
            checkout=parse_dt_24h(str(checkout_s)),
            meals_by_date=dict(meals_map),
            charge_room=(str(charge_room).upper() == "Y"),
        )

    def _group_by_vessel(self, stays: List[CrewStay]) -> Dict[str, List[CrewStay]]:
        groups: Dict[str, List[CrewStay]] = {}
        for s in stays:
            groups.setdefault(s.vessel, []).append(s)
        return groups

    def _generate_selected_group(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("안내", "생성할 행을 선택해 주세요.")
            return

        stays = [self._row_to_stay(iid) for iid in selected]
        groups = self._group_by_vessel(stays)

        created: List[Path] = []
        try:
            for vessel, rows in groups.items():
                created.append(save_invoice_by_vessel(vessel, rows, self.rates))
            self.status_var.set(f"선택 {len(created)}건(선박별) 생성 완료!")
            messagebox.showinfo("완료", "\n".join(str(p) for p in created))
        except Exception as e:
            _ui_error("생성 실패", f"{e}")
    def _generate_all_group(self):
        iids = self.tree.get_children()
        if not iids:
            messagebox.showinfo("안내", "목록이 비어 있습니다.")
            return

        stays = [self._row_to_stay(iid) for iid in iids]
        groups = self._group_by_vessel(stays)

        created: List[Path] = []
        try:
            for vessel, rows in groups.items():
                created.append(save_invoice_by_vessel(vessel, rows, self.rates))
            self.status_var.set(f"전체 {len(created)}건(선박별) 생성 완료!")
            messagebox.showinfo("완료", "\n".join(str(p) for p in created))
        except Exception as e:
            _ui_error("생성 실패", f"{e}")

    def _open_datetime_picker(self, target_var: tk.StringVar, title: str) -> None:
        """날짜는 타이핑 가능 + 달력 버튼(가능하면) + 시간 10분 단위 선택."""
        win = tk.Toplevel(self)
        win.title(title)
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)

        cur = (target_var.get() or "").strip()
        d0 = datetime.now().strftime("%Y-%m-%d")
        h0 = "10"
        m0 = "00"
        try:
            parts = cur.split()
            if len(parts) >= 1 and parts[0]:
                d0 = parts[0]
            if len(parts) >= 2 and ":" in parts[1]:
                h0, m0 = parts[1].split(":")[:2]
                h0 = h0.zfill(2)
                m0 = m0.zfill(2)
        except Exception:
            pass

        date_var = tk.StringVar(value=d0)
        hour_var = tk.StringVar(value=h0)
        min_var = tk.StringVar(value=m0)

        ttk.Label(frm, text="Date").grid(row=0, column=0, sticky="w")
        if False and DateEntry is not None:
            de = DateEntry(frm, textvariable=date_var, width=12, date_pattern="yyyy-mm-dd")
            de.grid(row=0, column=1, sticky="w", padx=(6, 0))
        else:
            ttk.Entry(frm, textvariable=date_var, width=12).grid(row=0, column=1, sticky="w", padx=(6, 0))

        ttk.Label(frm, text="Time").grid(row=1, column=0, sticky="w", pady=(8, 0))
        hours = [str(i).zfill(2) for i in range(24)]
        mins = ["00", "10", "20", "30", "40", "50"]

        cb_h = ttk.Combobox(frm, textvariable=hour_var, values=hours, width=4)
        cb_h.grid(row=1, column=1, sticky="w", padx=(6, 0), pady=(8, 0))
        cb_m = ttk.Combobox(frm, textvariable=min_var, values=mins, width=4)
        cb_m.grid(row=1, column=2, sticky="w", padx=(6, 0), pady=(8, 0))

        btns = ttk.Frame(frm)
        btns.grid(row=2, column=0, columnspan=3, sticky="e", pady=(12, 0))

        def ok():
            d = date_var.get().strip()
            hh = hour_var.get().strip().zfill(2)
            mm = min_var.get().strip().zfill(2)
            target_var.set(f"{d} {hh}:{mm}")
            win.destroy()

        ttk.Button(btns, text="Cancel", command=win.destroy).pack(side="right")
        ttk.Button(btns, text="OK", command=ok).pack(side="right", padx=(0, 6))


if __name__ == "__main__":
    App().mainloop()
